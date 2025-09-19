# %%

import os
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime
import time
import pyfiglet
import sys
import logging
import json
from difflib import get_close_matches

def setup_logging():
    """Set up comprehensive logging with file and console output"""
    # Create logs directory if it doesn't exist
    os.makedirs('logs', exist_ok=True)
    
    # Generate timestamp for log filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_filename = f'logs/flash_report_{timestamp}.log'
    
    # Configure logging with both file and console handlers
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    logger = logging.getLogger(__name__)
    logger.info("Flash Report logging has been successfully initialized.")
    logger.info(f"Log file created: {log_filename}")
    return logger

def print_banner(logger):
    now = datetime.now().strftime('%A, %d %B %Y %I:%M:%S %p')
    banner = pyfiglet.figlet_format(" Flash Report", font="standard")

    print(banner)  # Keep banner print for visual effect
    logger.info("Flash Report - Version v1.0.0")
    logger.info(f"Executed on: {now}")
    logger.info("Description: Comprehensive analysis and reporting tool initialized.")
    logger.info("=========================================")

# Initialize logging
logger = setup_logging()
print_banner(logger)


# %%
def exit_with_usage():
    """ Exit the program displaying usage information """
    logger.info("="*80)
    logger.info("FLASH REPORT TOOL - USAGE INFORMATION")
    logger.info("="*80)
    print("\nThis tool requires seven Excel files with specific formats:")
    print("\n1. MAIN DATA FILE (Web Intelligence Export):")
    print("   << Validation Checks >> ")
    print("   - SRC_SYS_KY, CROSS_SOURCED, BDE_FLAG, MSP_FLAG, REPORTING_TYPE")
    print("   - PRODUCT_LINE, RESELLER_PARTY_ID, DISTRIBUTOR_PARTY_ID, FISCAL_MONTH")
    print("   - NDP_TOTAL_USD, NET_TOTAL_USD, UPFRONT_DISCOUNT_AMT_USD, BACKEND_DISCOUNT_AMT_USD")
    print("   - Data Type, BACKEND_DEAL_1, Invoice Number, Hpe Sales Order Number")
    print("   - NET_TOTAL_LC, BACKEND_DISCOUNT_AMT_LC (for CA processing)")
    print("\n2. CA REFERENCE FILE:")
    print("   << Validation Checks >> ")
    print("   - PL, BU, Type, Exclusion_Party ID, Exclusion_Level")
    print("   - PG Exclusion Eligible List_Party ID, Loc Id, Elicpes")
    print("   - PN PL, BU.1, Common PL, Common PN PL")
    print("\n3. US REFERENCE FILE:")
    print("   << Validation Checks >> (Same as CA reference file)")
    print("   - PL, BU, Type, Exclusion_Party ID, Exclusion_Level")
    print("   - PG Exclusion Eligible List_Party ID, Loc Id, Elicpes")
    print("   - PN PL, BU.1, Common PL, Common PN PL")
    print("\n4. CA DAYS REPORTING FILE:")
    print("   << Format Requirements >> ")
    print("   - Must have at least 2 columns (date and days)")
    print("   - Date column should contain dates matching current execution date")
    print("\n5. US DAYS REPORTING FILE:")
    print("   << Format Requirements >> (Same as CA days reporting file)")
    print("   - Must have at least 2 columns (date and days)")
    print("   - Date column should contain dates matching current execution date")
    print("\n6. S3 Raw Data File (Operation 2):")
    print("   << All Sheets Reading >> ")
    print("   - Can contain multiple sheets")
    print("   - All sheets will be read and analyzed")
    print("\nFile Format Requirements:")
    print("   - All files must be Excel format (.xlsx)")
    print("   - Reference files must have data in 'Sheet1'")
    print("   - Column names must match exactly (case-sensitive)")
    print("Usage:")
    print("   python Flash_Report.py <main_data_file> <ca_reference_file> <us_reference_file> <ca_days_reporting_file> <us_days_reporting_file> <s3_raw_data> <rebate_file>")
    print("\nExample:")
    print("   python Flash_Report.py C:/data/Web_Intelligence_08072025.xlsx C:/ref/Flash_CA_ref_file.xlsx C:/ref/Flash_US_ref_file.xlsx C:/days/CA_days_reporting.xlsx C:/days/US_days_reporting.xlsx C:/days/s3_raw_data.xlsx C:/rebate/rebate_data.xlsx")
    print("\n" + "="*80)
    exit(1)

if len(sys.argv) != 8:
    print("Error: Exactly seven arguments required: main data file path, CA reference file path, US reference file path, CA FinBen reporting file path, US FinBen reporting file path, s3 raw data file path, rebate file path")
    exit_with_usage()

file_path = sys.argv[1]
source_path_ca = sys.argv[2]
source_path_us = sys.argv[3]
days_reporting_file_ca = sys.argv[4]
days_reporting_file_us = sys.argv[5]
s3_raw_data_path = sys.argv[6]
rebate_file_path = sys.argv[7]

# Validate the existence of files
for path, desc in [(file_path, 'main data file'), (source_path_ca, 'CA reference file'), (source_path_us, 'US reference file'), (days_reporting_file_ca, 'CA days reporting file'), (days_reporting_file_us, 'US days reporting file'), (s3_raw_data_path, 's3 raw data file'), (rebate_file_path, 'rebate file')]:
    if not os.path.isfile(path):
        print(f"Error: The {desc} at '{path}' does not exist.")
        exit_with_usage()

# %%
# Read the main file and validate format

def load_column_mappings():
    """ Load column mappings from JSON configuration file """
    try:
        with open('column_mappings.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        print("[WARNING] column_mappings.json not found, using default mappings")
        return get_default_mappings()
    except json.JSONDecodeError:
        print("[ERROR] Invalid JSON in column_mappings.json, using default mappings")
        return get_default_mappings()

def get_default_mappings():
    """ Fallback default mappings if JSON file is not available """
    return {
        "main_file_mappings": {
            "SRC_SYS_KY": ["SRC_SYS_KY", "Src Sys Ky"],
            "CROSS_SOURCED": ["CROSS_SOURCED", "Cross Sourced"],
            "BDE_FLAG": ["BDE_FLAG", "Bde Flag"],
            "MSP_FLAG": ["MSP_FLAG", "MSP Flag"],
            "REPORTING_TYPE": ["REPORTING_TYPE", "Reporting Type"],
            "PRODUCT_LINE": ["PRODUCT_LINE", "Product Line"],
            "RESELLER_PARTY_ID": ["RESELLER_PARTY_ID", "Reseller Party Id"],
            "DISTRIBUTOR_PARTY_ID": ["DISTRIBUTOR_PARTY_ID", "Distributor Party Id"],
            "FISCAL_MONTH": ["FISCAL_MONTH", "Fiscal Month"],
            "NDP_TOTAL_USD": ["NDP_TOTAL_USD", "Ndp Total Usd"],
            "NET_TOTAL_USD": ["NET_TOTAL_USD", "Net Total Usd"],
            "UPFRONT_DISCOUNT_AMT_USD": ["UPFRONT_DISCOUNT_AMT_USD", "Upfront Discount Amt Usd"],
            "BACKEND_DISCOUNT_AMT_USD": ["BACKEND_DISCOUNT_AMT_USD", "Backend Discount Amt Usd"],
            "DATA_TYPE": ["DATA_TYPE", "Data Type"],
            "BACKEND_DEAL_1": ["BACKEND_DEAL_1", "Backend Deal 1"],
            "INVOICE_NUMBER": ["INVOICE_NUMBER", "Invoice Number"],
            "HPE_SALES_ORDER_NUMBER": ["HPE_SALES_ORDER_NUMBER", "Hpe Sales Order Number"],
            "NET_TOTAL_LC": ["NET_TOTAL_LC", "Net Total Lc"],
            "BACKEND_DISCOUNT_AMT_LC": ["BACKEND_DISCOUNT_AMT_LC", "Backend Discount Amt Lc"],
            "UPFRONT_DISCOUNT_AMT_LC": ["UPFRONT_DISCOUNT_AMT_LC", "Upfront Discount Amt Lc"],
            "NDP_TOTAL_LC": ["NDP_TOTAL_LC", "Ndp Total Lc"],
            "DISTRIBUTOR_PARTY_NAME": ["Distributor Party Name", "DISTRIBUTOR_PARTY_NAME"],
            "RESELLER_PARTY_NAME": ["Reseller Party Name", "RESELLER_PARTY_NAME"],
            "PRODUCT_NUMBER": ["Product Number", "PRODUCT_NUMBER"]
        }
    }

def find_column_match(target_column, available_columns, mappings):
    """ Find the best match for a target column in available columns """
    if target_column in mappings:
        # Check exact matches first
        for variant in mappings[target_column]:
            if variant in available_columns:
                return variant
        
        # If no exact match, try fuzzy matching
        all_variants = mappings[target_column]
        for variant in all_variants:
            close_matches = get_close_matches(variant, available_columns, n=1, cutoff=0.8)
            if close_matches:
                return close_matches[0]
    
    # Last resort: direct fuzzy match on target column
    close_matches = get_close_matches(target_column, available_columns, n=1, cutoff=0.7)
    if close_matches:
        return close_matches[0]
    
    return None

def standardize_column_names(df, file_type='main'):
    """ Standardize column names using mappings and return column mapping dict """
    mappings = load_column_mappings()
    
    if file_type == 'main':
        target_mappings = mappings.get('main_file_mappings', {})
    else:
        target_mappings = mappings.get('reference_file_mappings', {})
    
    column_mapping = {}
    available_columns = list(df.columns)
    
    print(f"[*] Standardizing {file_type} file columns...")
    print(f"[*] Available columns: {available_columns[:10]}{'...' if len(available_columns) > 10 else ''}")
    
    # Find matches for each target column
    for target_col, variants in target_mappings.items():
        matched_col = find_column_match(target_col, available_columns, target_mappings)
        if matched_col:
            column_mapping[matched_col] = target_col
            print(f"[*] Mapped '{matched_col}' -> '{target_col}'")
        else:
            print(f"[WARNING] No match found for required column: {target_col}")
    
    # Rename columns in dataframe
    df_standardized = df.rename(columns=column_mapping)
    
    return df_standardized, column_mapping

def validate_main_file(df):
    """ Validate main data file format with flexible column matching """
    # Standardize column names first
    df_standardized, column_mapping = standardize_column_names(df, 'main')
    
    required_columns = ['SRC_SYS_KY', 'CROSS_SOURCED', 'BDE_FLAG', 'MSP_FLAG', 'REPORTING_TYPE', 
                       'PRODUCT_LINE', 'RESELLER_PARTY_ID', 'DISTRIBUTOR_PARTY_ID', 'FISCAL_MONTH',
                       'NDP_TOTAL_USD', 'NET_TOTAL_USD', 'UPFRONT_DISCOUNT_AMT_USD', 'BACKEND_DISCOUNT_AMT_USD',
                       'DATA_TYPE', 'BACKEND_DEAL_1', 'INVOICE_NUMBER', 'HPE_SALES_ORDER_NUMBER',
                       'NET_TOTAL_LC', 'BACKEND_DISCOUNT_AMT_LC', 'UPFRONT_DISCOUNT_AMT_LC', 'NDP_TOTAL_LC']
    
    # Check for critical missing columns
    missing_columns = [col for col in required_columns if col not in df_standardized.columns]
    
    if missing_columns:
        print(f"[ERROR] Main file is missing critical columns: {missing_columns}")
        print(f"[INFO] Mapped columns: {list(column_mapping.values())}")
        print(f"[INFO] Available columns: {list(df_standardized.columns)}")
        
        # Suggest possible matches for missing columns
        print("\n[SUGGESTIONS] Possible matches for missing columns:")
        for missing_col in missing_columns:
            suggestions = get_close_matches(missing_col, df.columns, n=3, cutoff=0.5)
            if suggestions:
                print(f"  {missing_col}: {suggestions}")
        
        exit_with_usage()
    
    print('[*] Main file format validation passed')
    return df_standardized

def validate_reference_file(df, file_type):
    """ Validate reference file format with flexible column matching """
    # Standardize column names first
    df_standardized, column_mapping = standardize_column_names(df, 'reference')
    
    required_columns = ['PL', 'BU', 'TYPE', 'EXCLUSION_PARTY_ID', 'EXCLUSION_LEVEL', 
                       'PG_EXCLUSION_ELIGIBLE_LIST_PARTY_ID', 'LOC_ID', 'ELICPES', 'PN_PL', 'BU_1', 
                       'COMMON_PL', 'COMMON_PN_PL']
    
    missing_columns = [col for col in required_columns if col not in df_standardized.columns]
    if missing_columns:
        print(f"[ERROR] {file_type} reference file is missing critical columns: {missing_columns}")
        print(f"[INFO] Mapped columns: {list(column_mapping.values())}")
        print(f"[INFO] Available columns: {list(df_standardized.columns)}")
        
        # Suggest possible matches for missing columns
        print(f"\n[SUGGESTIONS] Possible matches for missing columns in {file_type}:")
        for missing_col in missing_columns:
            suggestions = get_close_matches(missing_col, df.columns, n=3, cutoff=0.5)
            if suggestions:
                print(f"  {missing_col}: {suggestions}")
        
        exit_with_usage()
    
    print(f'[*] {file_type} reference file format validation passed')
    return df_standardized

logger.info("Starting input file validation process.")

try:
    df = pd.read_excel(file_path, engine='openpyxl')
    df = validate_main_file(df)
    logger.info(f"Main data file successfully loaded. Total rows: {len(df)}")
except Exception as e:
    print(f"Error reading main file: {e}")
    exit_with_usage()

try:
    print('[*] Reading the Reference File for CA')
    df_source_ca = pd.read_excel(source_path_ca, sheet_name='Sheet1')
    df_source_ca = validate_reference_file(df_source_ca, 'CA')
except Exception as e:
    print(f"Error reading CA reference file: {e}")
    exit_with_usage()

try:
    print('[*] Reading the Reference File for US')
    df_source_us = pd.read_excel(source_path_us, sheet_name='Sheet1')
    df_source_us = validate_reference_file(df_source_us, 'US')
except Exception as e:
    print(f"Error reading US reference file: {e}")
    exit_with_usage()

logger.info("All input files have been successfully loaded and validated.")

# Read and process Days of Reporting files
print('[*] Reading Days of Reporting files...')
today = datetime.today()
current_date = today.strftime('%Y-%m-%d')

# Read CA Days of Reporting file
try:
    print(f'[*] Reading CA Days of Reporting file: {days_reporting_file_ca}')
    df_days_ca = pd.read_excel(days_reporting_file_ca, engine='openpyxl')
    print(f'[*] CA Days of Reporting file loaded with {len(df_days_ca)} rows')
    
    # Get column names from the CA file
    ca_columns = df_days_ca.columns.tolist()
    if len(ca_columns) >= 2:
        date_col_ca = ca_columns[0]
        days_col_ca = ca_columns[1]
        
        # Convert date column to datetime for matching
        df_days_ca[date_col_ca] = pd.to_datetime(df_days_ca[date_col_ca], errors='coerce')
        current_date_dt = pd.to_datetime(current_date)
        
        # Find matching row for current date
        matching_row_ca = df_days_ca[df_days_ca[date_col_ca] == current_date_dt]
        
        if len(matching_row_ca) > 0:
            days_reporting_ca = int(matching_row_ca[days_col_ca].iloc[0])
            print(f'[*] CA Days of Reporting for {current_date}: {days_reporting_ca}')
        else:
            print(f'[!] ERROR: No exact match found for current date ({current_date}) in CA Days of Reporting file.')
            print(f'[!] The script must be executed on a valid reporting date that exists in the CA Days of Reporting file.')
            print(f'[!] Please check the CA Days of Reporting file and run the script on a valid date.')
            exit(1)
    else:
        print(f'[!] CA Days of Reporting file must have at least 2 columns (date and days)')
        days_reporting_ca = 0
        
except Exception as e:
    print(f'Error reading CA Days of Reporting file: {e}')
    days_reporting_ca = 0

# Read US Days of Reporting file
try:
    print(f'[*] Reading US Days of Reporting file: {days_reporting_file_us}')
    df_days_us = pd.read_excel(days_reporting_file_us, engine='openpyxl')
    print(f'[*] US Days of Reporting file loaded with {len(df_days_us)} rows')
    
    # Get column names from the US file
    us_columns = df_days_us.columns.tolist()
    if len(us_columns) >= 2:
        date_col_us = us_columns[0]
        days_col_us = us_columns[1]
        
        # Convert date column to datetime for matching
        df_days_us[date_col_us] = pd.to_datetime(df_days_us[date_col_us], errors='coerce')
        current_date_dt = pd.to_datetime(current_date)
        
        # Find matching row for current date
        matching_row_us = df_days_us[df_days_us[date_col_us] == current_date_dt]
        
        if len(matching_row_us) > 0:
            days_reporting_us = int(matching_row_us[days_col_us].iloc[0])
            print(f'[*] US Days of Reporting for {current_date}: {days_reporting_us}')
        else:
            print(f'[!] ERROR: No exact match found for current date ({current_date}) in US Days of Reporting file.')
            print(f'[!] The script must be executed on a valid reporting date that exists in the US Days of Reporting file.')
            print(f'[!] Please check the US Days of Reporting file and run the script on a valid date.')
            exit(1)
    else:
        print(f'[!] US Days of Reporting file must have at least 2 columns (date and days)')
        days_reporting_us = 0
        
except Exception as e:
    print(f'Error reading US Days of Reporting file: {e}')
    days_reporting_us = 0

print(f'[*] Days of Reporting loaded - CA: {days_reporting_ca}, US: {days_reporting_us}')

# ========================================================================================
# OPERATION 1: READ BOE Data, Validate and Filter to get the final data to DSO datasheet
# ========================================================================================

# %%
logger.info("Initiating data validation and filtering procedures.")

# Include rows with Src_sys_key == 2032 or 2866
df_src = df[df['SRC_SYS_KY'].isin([2032, 2866])] # type: ignore
print(f'[*] After filtering for SRC_SYS_KY: {len(df_src)} rows')


# %%
# Remove the rows with value 'Y' in the column
df_cross_sourced = df_src[df_src['CROSS_SOURCED'] != 'Y']
print(f'[*] After removing CROSS_SOURCED Y: {len(df_cross_sourced)} rows')

# %%
# Remove the 'Y' from BDE_FLAG and populate the blanks with 'N'
df_bde= df_cross_sourced[df_cross_sourced['BDE_FLAG'] != 'Y']
print(f'[*] After removing BDE_FLAG Y: {len(df_bde)} rows')
df_bde['BDE_FLAG'] = df_bde['BDE_FLAG'].fillna('N')

# %%
# Remove rows with values 'T' in MSP_FLAG column

df_msp = df_bde[df_bde['MSP_FLAG'] != 'T']
print(f'[*] After removing MSP_FLAG T: {len(df_msp)} rows')

# %%
# Remove rows with 'RCS' in 'REPORTING_TYPE' column

df_reporting = df_msp[df_msp['REPORTING_TYPE'] != 'RCS']
print(f'[*] After removing REPORTING_TYPE RCS: {len(df_reporting)} rows')

# Keeping the RCS data for future refrence in LA Sales
df_rcs = df_msp[df_msp['REPORTING_TYPE'] == 'RCS']
print(f'[*] RCS Data for reference: {len(df_rcs)} rows')

print('[*] Done!\n')
print('[*] Starting the additional column procedure...')

# %%
# Adding BU, BU_Types and Scheme_Name as columns
df_extend_columns = df_reporting.assign(BU='', BU_Type='',Scheme_Name='')
print(f'[*] After adding BU, BU_Type, Scheme_Name columns: {len(df_extend_columns)} rows')


# %%
# Reset the Index of the data
df_extend_columns = df_extend_columns.reset_index(drop=True)
print(f'[*] After resetting index: {len(df_extend_columns)} rows')

logger.info("Beginning region-specific data processing: US and CA paths.")

# %%
# ============================================================================
# US DATA PROCESSING PATH
# ============================================================================
print('[*] Processing US data with US reference file...')

# Create US dataframe copy
df_extend_columns_us = df_extend_columns.copy()
print(f'[*] Created US data copy: {len(df_extend_columns_us)} rows')

# Check if the PRODUCT_LINE value is in the 'PL' column of the US reference file
df_mapping_us = df_source_us.groupby('PL', as_index=True)[['BU', 'TYPE']].first() # type: ignore

df_extend_columns_us['BU'] = df_extend_columns_us['PRODUCT_LINE'].map(df_mapping_us['BU'])
df_extend_columns_us['BU_Type'] = df_extend_columns_us['PRODUCT_LINE'].map(df_mapping_us['TYPE'])
df_extend_columns_us['Scheme_Name'] = df_extend_columns_us['BU'].fillna('') + df_extend_columns_us['BU_Type'].fillna('')


# %%
# ============================================================================
# CA DATA PROCESSING PATH  
# ============================================================================
print('[*] Processing CA data with CA reference file...')

# Create CA dataframe copy
df_extend_columns_ca = df_extend_columns.copy()
print(f'[*] Created CA data copy: {len(df_extend_columns_ca)} rows')

# Check if the PRODUCT_LINE value is in the 'PL' column of the CA reference file
df_mapping_ca = df_source_ca.groupby('PL', as_index=True)[['BU', 'TYPE']].first() # type: ignore

df_extend_columns_ca['BU'] = df_extend_columns_ca['PRODUCT_LINE'].map(df_mapping_ca['BU'])
df_extend_columns_ca['BU_Type'] = df_extend_columns_ca['PRODUCT_LINE'].map(df_mapping_ca['TYPE'])
df_extend_columns_ca['Scheme_Name'] = df_extend_columns_ca['BU'].fillna('') + df_extend_columns_ca['BU_Type'].fillna('')


# %%
# ============================================================================
# US EXCLUSIONS PROCESSING
# ============================================================================
print('[*] Processing US exclusions and partner data...')

# Adding Exclusions, PG_Exclusions and Disty_Partners for US
df_exclusions_columns_us = df_extend_columns_us.assign(Exclusions='', PG_Exclusions='',Disty_Partners='')
print(f'[*] After adding exclusion columns to US: {len(df_exclusions_columns_us)} rows')

# If the RESELLER_PARTY_ID is in the Exclusion_Party_ID column then check 
# the column Exclusion_Level and populate the Exclusion Column
df_mapping_exc_us = df_source_us.groupby('EXCLUSION_PARTY_ID', as_index=True)['EXCLUSION_LEVEL'].first() # type: ignore
df_exclusions_columns_us['Exclusions'] = df_exclusions_columns_us['RESELLER_PARTY_ID'].map(df_mapping_exc_us)

# If the RESELLER_PARTY_ID (rpi) is in the PG Exclusion Eligible List_Party ID column 
# then insert PG in the PG_Exclusions column else SBP in the column.
df_exclusions_columns_us['PG_Exclusions'] = np.where(df_exclusions_columns_us['RESELLER_PARTY_ID'].isin(df_source_us['PG_EXCLUSION_ELIGIBLE_LIST_PARTY_ID']), 'PG', 'SBP') # type: ignore

# If the DISTRIBUTOR_PARTY_ID (dpi) is in the Loc Id column 
# then include the lookup value in the Disty_Partners column
try:
    df_exclusions_columns_us['Disty_Partners'] = np.where(df_exclusions_columns_us['DISTRIBUTOR_PARTY_ID'].isin(df_source_us['LOC_ID']), df_exclusions_columns_us['DISTRIBUTOR_PARTY_ID'], '') # type: ignore
except Exception as e:
    print(f"US - An unexpected error occurred: {e}")

# Converting the Disty_Partners column to float for US
df_exclusions_columns_us = df_exclusions_columns_us[df_exclusions_columns_us['Disty_Partners'] != '']
try:
    df_exclusions_columns_us['Disty_Partners'] = pd.to_numeric(df_exclusions_columns_us['Disty_Partners'], errors='coerce')
    df_exclusions_columns_us = df_exclusions_columns_us.dropna(subset=['Disty_Partners'])
except Exception as e:
    print(f"US - Error converting Disty_Partners to numeric: {e}")


# %%
# ============================================================================
# CA EXCLUSIONS PROCESSING
# ============================================================================
print('[*] Processing CA exclusions and partner data...')

# Adding Exclusions, PG_Exclusions and Disty_Partners for CA
df_exclusions_columns_ca = df_extend_columns_ca.assign(Exclusions='', PG_Exclusions='',Disty_Partners='')
print(f'[*] After adding exclusion columns to CA: {len(df_exclusions_columns_ca)} rows')

# If the RESELLER_PARTY_ID is in the Exclusion_Party_ID column then check 
# the column Exclusion_Level and populate the Exclusion Column
df_mapping_exc_ca = df_source_ca.groupby('EXCLUSION_PARTY_ID', as_index=True)['EXCLUSION_LEVEL'].first() # type: ignore
df_exclusions_columns_ca['Exclusions'] = df_exclusions_columns_ca['RESELLER_PARTY_ID'].map(df_mapping_exc_ca)

# If the RESELLER_PARTY_ID (rpi) is in the PG Exclusion Eligible List_Party ID column 
# then insert PG in the PG_Exclusions column else SBP in the column.
df_exclusions_columns_ca['PG_Exclusions'] = np.where(df_exclusions_columns_ca['RESELLER_PARTY_ID'].isin(df_source_ca['PG_EXCLUSION_ELIGIBLE_LIST_PARTY_ID']), 'PG', 'SBP') # type: ignore

# If the DISTRIBUTOR_PARTY_ID (dpi) is in the Loc Id column 
# then include the lookup value in the Disty_Partners column
try:
    df_exclusions_columns_ca['Disty_Partners'] = np.where(df_exclusions_columns_ca['DISTRIBUTOR_PARTY_ID'].isin(df_source_ca['LOC_ID']), df_exclusions_columns_ca['DISTRIBUTOR_PARTY_ID'], '') # type: ignore
except Exception as e:
    print(f"CA - An unexpected error occurred: {e}")

# Converting the Disty_Partners column to float for CA
df_exclusions_columns_ca = df_exclusions_columns_ca[df_exclusions_columns_ca['Disty_Partners'] != '']
try:
    df_exclusions_columns_ca['Disty_Partners'] = pd.to_numeric(df_exclusions_columns_ca['Disty_Partners'], errors='coerce')
    df_exclusions_columns_ca = df_exclusions_columns_ca.dropna(subset=['Disty_Partners'])
except Exception as e:
    print(f"CA - Error converting Disty_Partners to numeric: {e}")


# %%
# ============================================================================
# US CALCULATION COLUMNS PROCESSING
# ============================================================================
print('\n[*] Starting US Calculation of Metrics...')

# Adding the computation columns in the US formatted data table
df_exclusions_columns_calc_us = df_exclusions_columns_us.assign(Delta='', Updated_upfront='',Diff='', Match='', Match_1='')
print(f'[*] After adding calculation columns to US: {len(df_exclusions_columns_calc_us)} rows')

# Convert the columns to respective float datatype
df_exclusions_columns_calc_us['Delta'] = pd.to_numeric(df_exclusions_columns_calc_us['Delta'], errors='coerce')
df_exclusions_columns_calc_us['Updated_upfront'] = pd.to_numeric(df_exclusions_columns_calc_us['Updated_upfront'], errors='coerce')
df_exclusions_columns_calc_us['Diff'] = pd.to_numeric(df_exclusions_columns_calc_us['Diff'], errors='coerce')
df_exclusions_columns_calc_us['Match'] = pd.to_numeric(df_exclusions_columns_calc_us['Match'], errors='coerce')
df_exclusions_columns_calc_us['Match_1'] = pd.to_numeric(df_exclusions_columns_calc_us['Match_1'], errors='coerce')

# US Calculations
df_exclusions_columns_calc_us['Delta'] = (df_exclusions_columns_calc_us['NDP_TOTAL_USD'] - df_exclusions_columns_calc_us['UPFRONT_DISCOUNT_AMT_USD'] - df_exclusions_columns_calc_us['BACKEND_DISCOUNT_AMT_USD']) - df_exclusions_columns_calc_us['NET_TOTAL_USD']
df_exclusions_columns_calc_us['Updated_upfront'] = (df_exclusions_columns_calc_us['Delta'] + df_exclusions_columns_calc_us['UPFRONT_DISCOUNT_AMT_USD'])
df_exclusions_columns_calc_us['Diff'] = (df_exclusions_columns_calc_us['NDP_TOTAL_USD'] - df_exclusions_columns_calc_us['BACKEND_DISCOUNT_AMT_USD'] - df_exclusions_columns_calc_us['Updated_upfront'] - df_exclusions_columns_calc_us['NET_TOTAL_USD'])
df_exclusions_columns_calc_us['Match'] = df_exclusions_columns_calc_us['NDP_TOTAL_USD'] -(df_exclusions_columns_calc_us['Updated_upfront'] + df_exclusions_columns_calc_us['BACKEND_DISCOUNT_AMT_USD'])
df_exclusions_columns_calc_us['Match_1'] = df_exclusions_columns_calc_us['Match'] - df_exclusions_columns_calc_us['NET_TOTAL_USD']


# %%
# ============================================================================
# CA CALCULATION COLUMNS PROCESSING
# ============================================================================
print('[*] Starting CA Calculation of Metrics...')

# Adding the computation columns in the CA formatted data table
df_exclusions_columns_calc_ca = df_exclusions_columns_ca.assign(Delta='', Updated_upfront='',Diff='', Match='', Match_1='')
print(f'[*] After adding calculation columns to CA: {len(df_exclusions_columns_calc_ca)} rows')

# Convert the columns to respective float datatype
df_exclusions_columns_calc_ca['Delta'] = pd.to_numeric(df_exclusions_columns_calc_ca['Delta'], errors='coerce')
df_exclusions_columns_calc_ca['Updated_upfront'] = pd.to_numeric(df_exclusions_columns_calc_ca['Updated_upfront'], errors='coerce')
df_exclusions_columns_calc_ca['Diff'] = pd.to_numeric(df_exclusions_columns_calc_ca['Diff'], errors='coerce')
df_exclusions_columns_calc_ca['Match'] = pd.to_numeric(df_exclusions_columns_calc_ca['Match'], errors='coerce')
df_exclusions_columns_calc_ca['Match_1'] = pd.to_numeric(df_exclusions_columns_calc_ca['Match_1'], errors='coerce')

# CA Calculations
df_exclusions_columns_calc_ca['Delta'] = (df_exclusions_columns_calc_ca['NDP_TOTAL_LC'] - df_exclusions_columns_calc_ca['UPFRONT_DISCOUNT_AMT_LC'] - df_exclusions_columns_calc_ca['BACKEND_DISCOUNT_AMT_LC']) - df_exclusions_columns_calc_ca['NET_TOTAL_LC']
df_exclusions_columns_calc_ca['Updated_upfront'] = (df_exclusions_columns_calc_ca['Delta'] + df_exclusions_columns_calc_ca['UPFRONT_DISCOUNT_AMT_LC'])
df_exclusions_columns_calc_ca['Diff'] = (df_exclusions_columns_calc_ca['NDP_TOTAL_LC'] - df_exclusions_columns_calc_ca['BACKEND_DISCOUNT_AMT_LC'] - df_exclusions_columns_calc_ca['Updated_upfront'] - df_exclusions_columns_calc_ca['NET_TOTAL_LC'])
df_exclusions_columns_calc_ca['Match'] = df_exclusions_columns_calc_ca['NDP_TOTAL_LC'] -(df_exclusions_columns_calc_ca['Updated_upfront'] + df_exclusions_columns_calc_ca['BACKEND_DISCOUNT_AMT_LC'])
df_exclusions_columns_calc_ca['Match_1'] = df_exclusions_columns_calc_ca['Match'] - df_exclusions_columns_calc_ca['NET_TOTAL_LC']


# %%
# ============================================================================
# US FINAL COLUMNS PROCESSING
# ============================================================================
print('[*] Processing US final columns (PIPP, PN_Standalone, Common_PN_PL)...')

# Adding PIPP_Delas, PN_Standalone and Common_PN_PL for US
df_exclusions_columns_final_us = df_exclusions_columns_calc_us.assign(PIPP_delas='', PN_Standalone='', Common_PN_PL='')
print(f'[*] After US final columns processing: {len(df_exclusions_columns_final_us)} rows')

# If the BACKEND_DEAL_1 is in the Elicpes column then include the value in the PIPP delas column
df_exclusions_columns_final_us['PIPP_delas'] = df_exclusions_columns_final_us['BACKEND_DEAL_1'].where(df_exclusions_columns_final_us['BACKEND_DEAL_1'].isin(df_source_us['ELICPES'])) # type: ignore

# If the PRODUCT_LINE is in the PN PL column then check the column BU and populate the PN_Standalone (pns) Column
df_mapping_pns_us = df_source_us.groupby('PN_PL', as_index=True)['BU_1'].first() # type: ignore
df_exclusions_columns_final_us['PN_Standalone'] = df_exclusions_columns_final_us['PRODUCT_LINE'].map(df_mapping_pns_us)

# If the PRODUCT_LINE is in the Common PL column then check the column Common PN PL column and populate the Common_PN_PL (cpp) Column
df_mapping_pnpl_us = df_source_us.groupby('COMMON_PL', as_index=True)['COMMON_PN_PL'].first() # type: ignore
df_exclusions_columns_final_us['Common_PN_PL'] = df_exclusions_columns_final_us['PRODUCT_LINE'].map(df_mapping_pnpl_us)


# %%
# ============================================================================
# CA FINAL COLUMNS PROCESSING
# ============================================================================
print('[*] Processing CA final columns (PIPP, PN_Standalone, Common_PN_PL)...')

# Adding PIPP_Delas, PN_Standalone and Common_PN_PL for CA
df_exclusions_columns_final_ca = df_exclusions_columns_calc_ca.assign(PIPP_delas='', PN_Standalone='', Common_PN_PL='')
print(f'[*] After CA final columns processing: {len(df_exclusions_columns_final_ca)} rows')

# If the BACKEND_DEAL_1 is in the Elicpes column then include the value in the PIPP delas column
df_exclusions_columns_final_ca['PIPP_delas'] = df_exclusions_columns_final_ca['BACKEND_DEAL_1'].where(df_exclusions_columns_final_ca['BACKEND_DEAL_1'].isin(df_source_ca['ELICPES'])) # type: ignore

# If the PRODUCT_LINE is in the PN PL column then check the column BU and populate the PN_Standalone (pns) Column
df_mapping_pns_ca = df_source_ca.groupby('PN_PL', as_index=True)['BU_1'].first() # type: ignore
df_exclusions_columns_final_ca['PN_Standalone'] = df_exclusions_columns_final_ca['PRODUCT_LINE'].map(df_mapping_pns_ca)

# If the PRODUCT_LINE is in the Common PL column then check the column Common PN PL column and populate the Common_PN_PL (cpp) Column
df_mapping_pnpl_ca = df_source_ca.groupby('COMMON_PL', as_index=True)['COMMON_PN_PL'].first() # type: ignore
df_exclusions_columns_final_ca['Common_PN_PL'] = df_exclusions_columns_final_ca['PRODUCT_LINE'].map(df_mapping_pnpl_ca)

# %%

# ============================================================================
# ADDITIONAL FORMATTING AND REPORT GENERATION
# ============================================================================
print('\n[*] Starting formatting of additional columns...')
time.sleep(2)
# Add further processing as needed for final report generation

# ============================================================================
# CREATE FORMATTED DATA COPIES FOR LATER USE
# ============================================================================
print('\n[*] Creating formatted data copies for monthly sales calculations...')

# Create formatted copy for US data
# Filter based on Scheme_Name: exclude empty values
# Filter PIPP_delas: consider only NaN values
df_exclusions_columns_final_us_formatted = df_exclusions_columns_final_us[     
    (df_exclusions_columns_final_us['Scheme_Name'] != '') & 
    (df_exclusions_columns_final_us['PIPP_delas'].isna())
].copy()

print(f'[*] Created US formatted data copy: {len(df_exclusions_columns_final_us_formatted)} rows')
print(f'    - Original US data: {len(df_exclusions_columns_final_us)} rows')
print(f'    - Filtered for valid Scheme_Name and NaN/blank PIPP_delas')

# Create formatted copy for CA data
# Filter based on Scheme_Name: exclude empty values
# Filter PIPP_delas: consider only NaN values
df_exclusions_columns_final_ca_formatted = df_exclusions_columns_final_ca[
    (df_exclusions_columns_final_ca['Scheme_Name'] != '') &
    (df_exclusions_columns_final_ca['PIPP_delas'].isna())
].copy()

print(f'[*] Created CA formatted data copy: {len(df_exclusions_columns_final_ca_formatted)} rows')
print(f'    - Original CA data: {len(df_exclusions_columns_final_ca)} rows')
print(f'    - Filtered for valid Scheme_Name and NaN/blank PIPP_delas')

# Create unfiltered formatted copies that INCLUDE exclusions for monthly sales calculation
# Use the data BEFORE exclusions filtering (df_exclusions_columns_final_us/ca contains ALL data including exclusions)
df_exclusions_columns_final_us_formatted_with_exclusions = df_exclusions_columns_final_us[
    (df_exclusions_columns_final_us['Scheme_Name'] != '') & 
    (df_exclusions_columns_final_us['PIPP_delas'].isna())
].copy()

df_exclusions_columns_final_ca_formatted_with_exclusions = df_exclusions_columns_final_ca[
    (df_exclusions_columns_final_ca['Scheme_Name'] != '') &
    (df_exclusions_columns_final_ca['PIPP_delas'].isna())
].copy()

print(f'[*] Created US formatted data with exclusions: {len(df_exclusions_columns_final_us_formatted_with_exclusions)} rows')
print(f'[*] Created CA formatted data with exclusions: {len(df_exclusions_columns_final_ca_formatted_with_exclusions)} rows')

print('[*] Formatted data copies created successfully for monthly sales processing.')

# %%

# Formatting additional columns for US and CA

# US Formatting
print('[*] Formatting US additional columns...')
###### US - BU ######
df_final_us = df_exclusions_columns_final_us.dropna(subset=['BU'])
print(f'[*] After formatting US BU columns: {len(df_final_us)} rows')

###### US - Exclusions #####
# Replace 'NA' with np.nan
df_final_us['Exclusions'] = df_final_us['Exclusions'].replace('NA', np.nan)
# Filter rows where 'Exclusions' is NaN or empty string
df_final_us_exclusion = df_final_us[~df_final_us['Exclusions'].isna()]
df_final_us = df_final_us[df_final_us['Exclusions'].isna()]

###### US - Disty_Partners #####
# Keep only rows with non-empty Disty_Partners
df_final_us = df_final_us[~(df_final_us['Disty_Partners'].isna() | (df_final_us['Disty_Partners'] == ''))]


# %%
# CA Formatting
print('[*] Formatting CA additional columns...')
###### CA - BU ######
df_final_ca = df_exclusions_columns_final_ca.dropna(subset=['BU'])
print(f'[*] After formatting CA BU columns: {len(df_final_ca)} rows')

###### CA - Exclusions #####
# Replace 'NA' with np.nan
df_final_ca['Exclusions'] = df_final_ca['Exclusions'].replace('NA', np.nan)
# Filter rows where 'Exclusions' is NaN or empty string
df_final_ca_exclusion = df_final_ca[~df_final_ca['Exclusions'].isna()]
df_final_ca = df_final_ca[df_final_ca['Exclusions'].isna()]

###### CA - Disty_Partners #####
# Keep only rows with non-empty Disty_Partners
df_final_ca = df_final_ca[~(df_final_ca['Disty_Partners'].isna() | (df_final_ca['Disty_Partners'] == ''))]


# %%
# ============================================================================
# PG/SBP SEPARATION FOR US AND CA
# ============================================================================
print('[*] Separating US data into PG and SBP categories...')
# US PG/SBP separation
df_pg_us = df_final_us[df_final_us['PG_Exclusions'] == 'PG']
print(f'[*] US PG rows: {len(df_pg_us)} rows')
df_sbp_us = df_final_us[df_final_us['PG_Exclusions'] == 'SBP']
print(f'[*] US SBP rows: {len(df_sbp_us)} rows')

print('[*] Separating CA data into PG and SBP categories...')
# CA PG/SBP separation
df_pg_ca = df_final_ca[df_final_ca['PG_Exclusions'] == 'PG']
print(f'[*] CA PG rows: {len(df_pg_ca)} rows')
df_sbp_ca = df_final_ca[df_final_ca['PG_Exclusions'] == 'SBP']
print(f'[*] CA SBP rows: {len(df_sbp_ca)} rows')

print(f'\n[*] US Dataset - Total rows after formatting: {len(df_final_us)}')
print(f'[*] US Dataset - PG rows: {len(df_pg_us)}, SBP rows: {len(df_sbp_us)}')
print(f'[*] CA Dataset - Total rows after formatting: {len(df_final_ca)}')
print(f'[*] CA Dataset - PG rows: {len(df_pg_ca)}, SBP rows: {len(df_sbp_ca)}')


# %%
# ============================================================================
# MONTHLY REPORT PREPARATION FOR US AND CA
# ============================================================================
print('\n[*] Preparing Monthly Reports for US and CA...')

def prepare_monthly_data(df_data, region_name):
    """
    Prepare monthly data for report generation
    """
    print(f'[*] Processing {region_name} monthly data...')
    
    # Convert 'month' from YYYYMM to datetime
    df_data['month_str'] = df_data['FISCAL_MONTH'].astype(str) + '01'
    df_data['month_date'] = pd.to_datetime(df_data['month_str'], format='%Y%m%d')
    
    # Define fiscal year quarters (starting from November)
    def get_fiscal_quarter(month):
        if month in [11, 12, 1]:
            return 'Q1'
        elif month in [2, 3, 4]:
            return 'Q2'
        elif month in [5, 6, 7]:
            return 'Q3'
        else:
            return 'Q4'
    
    df_data['fiscal_quarter'] = df_data['month_date'].dt.month.apply(get_fiscal_quarter)
    
    # Filter by current fiscal quarter
    today = datetime.today()
    current_month = today.month
    
    # Determine current fiscal quarter
    if current_month in [11, 12, 1]:
        current_fq = 'Q1'
    elif current_month in [2, 3, 4]:
        current_fq = 'Q2'
    elif current_month in [5, 6, 7]:
        current_fq = 'Q3'
    else:
        current_fq = 'Q4'
    
    # Apply the filter
    df_data = df_data[df_data['fiscal_quarter'] == current_fq]
    
    # Create month names
    df_data['month_name'] = df_data['month_date'].dt.strftime('%B')
    df_data['month_num'] = df_data['month_date'].dt.month
    df_data['month_sales_col'] = df_data['month_name'] + '_NDP_sales'
    
    print(f'[*] {region_name} data prepared - {len(df_data)} rows for current fiscal quarter')
    return df_data

# Prepare monthly data for both US and CA using data WITH exclusions for monthly sales
df_experiment_us = prepare_monthly_data(df_exclusions_columns_final_us_formatted_with_exclusions.copy(), 'US')
print(f'[*] Prepared US monthly data: {len(df_experiment_us)} rows')
df_experiment_ca = prepare_monthly_data(df_exclusions_columns_final_ca_formatted_with_exclusions.copy(), 'CA')
print(f'[*] Prepared CA monthly data: {len(df_experiment_ca)} rows')


# %%
# ============================================================================
# REPORT GENERATION FUNCTION FOR US AND CA
# ============================================================================
def generate_currency_report_regional(df_main_data, df_exclusion, df_pg, df_sbp, region_name, currency_type):
    """
    Generate report for USD or LC currency
    currency_type: 'Usd' or 'Lc'
    
    Note: df_pg and df_sbp should be df_final_us/df_final_ca for PG/SBP Coverage Sales
          df_data should be df_exclusions_columns_final_us_triggered/df_exclusions_columns_final_ca for other calculations
          
    IMPORTANT: Monthly sales columns will use ALL data (no exclusions filter)
               Other metrics will use exclusions-filtered data
    """
    print(f"[*] Generating {currency_type} report for {region_name} using formatted data ({len(df_main_data)} rows)...")
    
    # Create unfiltered data for monthly sales (includes exclusions)
    # We need to get the original data before exclusions filtering
    # The df_main_data passed in should be the formatted data that includes exclusions
    df_main_data_unfiltered = df_main_data.copy()

    # For Scheme_Name with 'Compute' or 'Storage': only include Data Type = 'DS'
    compute_storage_filtered_base = df_main_data[
        df_main_data['Scheme_Name'].str.contains('Compute|Storage', case=False, na=False) & 
        (df_main_data['DATA_TYPE'] == 'DS')
    ]
    
    # Rows with Scheme_Name containing Compute or Storage, Data Type 'Orders' or 'S4DOR' and PRODUCT_LINE 'N3', 
    # to be added to compute_storage_filtered instead
    compute_storage_orders_n3 = df_main_data[
        df_main_data['Scheme_Name'].str.contains('Compute|Storage', case=False, na=False) & 
        (df_main_data['DATA_TYPE'].isin(['Orders', 'S4DOR'])) & 
        (df_main_data['PRODUCT_LINE'] == 'N3')
    ]
    
    # Combine compute/storage DS data with these additional compute/storage 'Orders' rows
    compute_storage_filtered = pd.concat([compute_storage_filtered_base, compute_storage_orders_n3], ignore_index=True)
    
    # Services data remains separate
    services_filtered = df_main_data[
        df_main_data['Scheme_Name'].str.contains('Services', case=False, na=False) &
        (df_main_data['DATA_TYPE'].isin(['DS', 'Orders', 'S4DOR']))
    ]

    # Apply Services-specific Scheme_Name updates
    print(f"[*] Processing Services data for {region_name} - {len(services_filtered)} rows")
    
    # Track number of rows affected by changes
    rows_affected = 0
    
    # Create a copy to avoid SettingWithCopyWarning
    services_filtered = services_filtered.copy()
    
    # Iterate through services data to apply the transformation logic
    for idx in services_filtered.index:
        pn_standalone = services_filtered.loc[idx, 'PN_Standalone']
        
        # Keep rows where PN_Standalone is 'Services Focus standalone' or 'Services Standard standalone'
        if pn_standalone in ['Services Focus standalone', 'Services Standard standalone']:
            continue
        
        # For rows where PN_Standalone is empty or NaN
        if pd.isna(pn_standalone) or pn_standalone == '':
            common_pn_pl = services_filtered.loc[idx, 'Common_PN_PL']
            
            # Check if Common_PN_PL has a value and equals 'Common PL'
            if not (pd.isna(common_pn_pl) or common_pn_pl == '') and common_pn_pl == 'Common PL':
                # Get the portfolio mapping values
                portfolio_mapping_1 = services_filtered.loc[idx, 'Point Next Portfolio Mapping 1'] if 'Point Next Portfolio Mapping 1' in services_filtered.columns else ''
                portfolio_mapping_2 = services_filtered.loc[idx, 'Point Next Portfolio Mapping 2'] if 'Point Next Portfolio Mapping 2' in services_filtered.columns else ''
                
                # Apply transformation logic using "starts with" matching
                if (str(portfolio_mapping_1).startswith('Operational Service') and 
                    str(portfolio_mapping_2).startswith('Complete Care (excl. MS & GL)')):
                    services_filtered.loc[idx, 'Scheme_Name'] = 'ServicesFocus'
                    rows_affected += 1
                elif (not (pd.isna(portfolio_mapping_1) or portfolio_mapping_1 == '') and 
                      not (pd.isna(portfolio_mapping_2) or portfolio_mapping_2 == '') and # take everything except Complete Care (excl. MS & GL)
                      not str(portfolio_mapping_2).startswith('Complete Care (excl. MS & GL)')):
                    services_filtered.loc[idx, 'Scheme_Name'] = 'ServicesStandard'
                    rows_affected += 1
    
    print(f"[*] Services data processing completed - {rows_affected} rows had Scheme_Name updated")

    # Combine all filtered dataframes (for non-monthly calculations)
    df_combined_filtered = pd.concat([compute_storage_filtered, services_filtered], ignore_index=True)
    
    # Create unfiltered versions for monthly sales calculation
    # Apply same scheme filtering but without exclusions filter
    compute_storage_unfiltered_base = df_main_data_unfiltered[
        df_main_data_unfiltered['Scheme_Name'].str.contains('Compute|Storage', case=False, na=False) & 
        (df_main_data_unfiltered['DATA_TYPE'] == 'DS')
    ]
    
    compute_storage_unfiltered_orders_n3 = df_main_data_unfiltered[
        df_main_data_unfiltered['Scheme_Name'].str.contains('Compute|Storage', case=False, na=False) & 
        (df_main_data_unfiltered['DATA_TYPE'].isin(['Orders', 'S4DOR'])) & 
        (df_main_data_unfiltered['PRODUCT_LINE'] == 'N3')
    ]
    
    compute_storage_unfiltered = pd.concat([compute_storage_unfiltered_base, compute_storage_unfiltered_orders_n3], ignore_index=True)
    
    services_unfiltered = df_main_data_unfiltered[
        df_main_data_unfiltered['Scheme_Name'].str.contains('Services', case=False, na=False) &
        (df_main_data_unfiltered['DATA_TYPE'].isin(['DS', 'Orders', 'S4DOR']))
    ].copy()
    
    # Apply same Services-specific Scheme_Name updates to unfiltered data
    for idx in services_unfiltered.index:
        pn_standalone = services_unfiltered.loc[idx, 'PN_Standalone']
        
        if pn_standalone in ['Services Focus standalone', 'Services Standard standalone']:
            continue
        
        if pd.isna(pn_standalone) or pn_standalone == '':
            common_pn_pl = services_unfiltered.loc[idx, 'Common_PN_PL']
            
            if not (pd.isna(common_pn_pl) or common_pn_pl == '') and common_pn_pl == 'Common PL':
                portfolio_mapping_1 = services_unfiltered.loc[idx, 'Point Next Portfolio Mapping 1'] if 'Point Next Portfolio Mapping 1' in services_unfiltered.columns else ''
                portfolio_mapping_2 = services_unfiltered.loc[idx, 'Point Next Portfolio Mapping 2'] if 'Point Next Portfolio Mapping 2' in services_unfiltered.columns else ''
                
                if (str(portfolio_mapping_1).startswith('Operational Service') and 
                    str(portfolio_mapping_2).startswith('Complete Care (excl. MS & GL)')):
                    services_unfiltered.loc[idx, 'Scheme_Name'] = 'ServicesFocus'
                elif (not (pd.isna(portfolio_mapping_1) or portfolio_mapping_1 == '') and 
                      not (pd.isna(portfolio_mapping_2) or portfolio_mapping_2 == '') and 
                      not str(portfolio_mapping_2).startswith('Complete Care (excl. MS & GL)')):
                    services_unfiltered.loc[idx, 'Scheme_Name'] = 'ServicesStandard'
    
    # Combine unfiltered data for monthly sales pivot
    df_combined_unfiltered = pd.concat([compute_storage_unfiltered, services_unfiltered], ignore_index=True)
    
    print(f"[*] Using unfiltered data for monthly sales: {len(df_combined_unfiltered)} rows (includes exclusions)")
    print(f"[*] Using filtered data for other metrics: {len(df_combined_filtered)} rows (excludes exclusions)")

    # Create pivot table with appropriate currency column using UNFILTERED data for monthly sales
    df_report = df_combined_unfiltered.pivot_table(
        index=['Scheme_Name', 'PRODUCT_LINE'],
        columns='month_sales_col',
        values=f'NDP_TOTAL_{currency_type.upper()}',
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    # Rearrange columns (keep Scheme, PRODUCT_LINEs, then sales columns)
    # Use unfiltered data for month ordering since that's what we used for pivot
    month_order = df_combined_unfiltered[['month_sales_col', 'month_num']].drop_duplicates()
    month_order = month_order.sort_values('month_num')
    ordered_month_cols = month_order['month_sales_col'].tolist()

    # Final column order: scheme_name, product_line, then month columns in correct order
    df_report = df_report[['Scheme_Name', 'PRODUCT_LINE'] + ordered_month_cols]
    float_cols = df_report.select_dtypes(include='float').columns
    df_report[float_cols] = df_report[float_cols].round(2)

    # Computing Total Sales
    sales_cols = [col for col in df_report.columns if col.endswith('_sales')]
    df_report['Total sales'] = df_report[sales_cols].sum(axis=1)
    df_report['Total sales'] = df_report['Total sales'].round(2)

    # Adding the Updated_upfront and Backend column data into the report
    agg_df = df_combined_filtered.groupby(['Scheme_Name', 'PRODUCT_LINE'])[['Updated_upfront', f'BACKEND_DISCOUNT_AMT_{currency_type.upper()}']].sum().reset_index()

    # Debug: Check for duplicates in df_report and agg_df before merging
    print(f"[DEBUG] df_report duplicates count on key columns: {df_report.duplicated(subset=['Scheme_Name', 'PRODUCT_LINE']).sum()}")
    print(f"[DEBUG] agg_df duplicates count on key columns: {agg_df.duplicated(subset=['Scheme_Name', 'PRODUCT_LINE']).sum()}")

    # Remove duplicates if present to prevent InvalidIndexError
    if df_report.duplicated(subset=['Scheme_Name', 'PRODUCT_LINE']).any():
        original_count = len(df_report)
        print("[WARNING] Duplicates found in df_report - removing duplicates")
        df_report = df_report.drop_duplicates(subset=['Scheme_Name', 'PRODUCT_LINE'])
        removed_count = original_count - len(df_report)
        print(f"[INFO] Removed {removed_count} duplicate rows from df_report")
    if agg_df.duplicated(subset=['Scheme_Name', 'PRODUCT_LINE']).any():
        original_count = len(agg_df)
        print("[WARNING] Duplicates found in agg_df - removing duplicates")
        agg_df = agg_df.drop_duplicates(subset=['Scheme_Name', 'PRODUCT_LINE'])
        removed_count = original_count - len(agg_df)
        print(f"[INFO] Removed {removed_count} duplicate rows from agg_df")

    # Merge the Updated_upfronts into df_report
    df_report = df_report.merge(agg_df, on=['Scheme_Name', 'PRODUCT_LINE'], how='left')

    df_report[['Updated_upfront', f'BACKEND_DISCOUNT_AMT_{currency_type.upper()}']] = df_report[['Updated_upfront', f'BACKEND_DISCOUNT_AMT_{currency_type.upper()}']].round(2)

    # Renaming columns for convenience
    df_report.rename(columns={
        'Scheme_Name': 'Program Name',
        'Updated_upfront': 'Upfront',
        f'BACKEND_DISCOUNT_AMT_{currency_type.upper()}': 'Backend',
        'Total sales': 'NDP Sales TSO'
    }, inplace=True)

    # Total Deal @Net Sales Out
    df_report['Total Deal @Net Sales Out'] = df_report['NDP Sales TSO'] - (df_report['Upfront'] + df_report['Backend'])
    df_report = df_report.round(2)

    # Computing Exclusions Sales
    df_filtered_exc = df_exclusion[df_exclusion['Exclusions'] != 'OEM']
    other_exclusions_df = df_filtered_exc.groupby('PRODUCT_LINE')[f'NET_TOTAL_{currency_type.upper()}'].sum().reset_index()
    other_exclusions_df.rename(columns={f'NET_TOTAL_{currency_type.upper()}': 'Other Exclusions Net Sales'}, inplace=True)

    # Merge into df_report based on product_line
    df_report = df_report.merge(other_exclusions_df, on='PRODUCT_LINE', how='left')
    df_report['Other Exclusions Net Sales'] = df_report['Other Exclusions Net Sales'].fillna(0).round(2)

    # OEM Exclusions
    oem_df = df_exclusion[df_exclusion['Exclusions'] == 'OEM']
    oem_exclusions_df = oem_df.groupby('PRODUCT_LINE')[f'NET_TOTAL_{currency_type.upper()}'].sum().reset_index()
    oem_exclusions_df.rename(columns={f'NET_TOTAL_{currency_type.upper()}': 'OEM Exclusions'}, inplace=True)
    df_report = df_report.merge(oem_exclusions_df, on='PRODUCT_LINE', how='left')
    df_report['OEM Exclusions'] = df_report['OEM Exclusions'].fillna(0).round(2)

    # LA Sales Exclusions
    # based on the DISTRIBUTOR_PARTY_ID and PRODUCT_LINE
    rcs_exclusions_df = df_rcs.groupby(['PRODUCT_LINE', 'DISTRIBUTOR_PARTY_ID'])[f'NET_TOTAL_{currency_type.upper()}'].sum().reset_index()
    # Sum by PRODUCT_LINE for merging with df_report (which only has PRODUCT_LINE)
    rcs_exclusions_summary = rcs_exclusions_df.groupby('PRODUCT_LINE')['NET_TOTAL_{}'.format(currency_type.upper())].sum().reset_index()
    rcs_exclusions_summary.rename(columns={f'NET_TOTAL_{currency_type.upper()}': 'LA Sales'}, inplace=True)
    df_report = df_report.merge(rcs_exclusions_summary, on='PRODUCT_LINE', how='left')
    df_report['LA Sales'] = df_report['LA Sales'].fillna(0).round(2)

    # Total Exclusion Sales
    df_report['Total Exclusions'] = df_report['Other Exclusions Net Sales'] + df_report['OEM Exclusions'] + df_report['LA Sales']
    df_report['Total Exclusions'] = df_report['Total Exclusions'].fillna(0).round(2)

    # Total eligible sales (Total Sales - Total Exclusions)
    df_report['Total eligible DSO Deal@Net'] = df_report['Total Deal @Net Sales Out'] - df_report['Total Exclusions']
    df_report['Total eligible DSO Deal@Net'] = df_report['Total eligible DSO Deal@Net'].fillna(0).round(2)

    # Calculating the PG Coverage Sales and SBP Coverage Sales columns
    pg_exclusions_df = df_pg.groupby('PRODUCT_LINE')[f'NET_TOTAL_{currency_type.upper()}'].sum().reset_index()
    pg_exclusions_df.rename(columns={f'NET_TOTAL_{currency_type.upper()}': 'PG Coverage Sales'}, inplace=True)
    df_report = df_report.merge(pg_exclusions_df, on='PRODUCT_LINE', how='left')
    df_report['PG Coverage Sales'] = df_report['PG Coverage Sales'].fillna(0).round(2)

    sbp_exclusions_df = df_sbp.groupby('PRODUCT_LINE')[f'NET_TOTAL_{currency_type.upper()}'].sum().reset_index()
    sbp_exclusions_df.rename(columns={f'NET_TOTAL_{currency_type.upper()}': 'SBP Coverage Sales'}, inplace=True)
    df_report = df_report.merge(sbp_exclusions_df, on='PRODUCT_LINE', how='left')
    df_report['SBP Coverage Sales'] = df_report['SBP Coverage Sales'].fillna(0).round(2)

    return df_report

# Generate reports for both US and CA datasets
# Note: df_experiment_us/ca are passed but function uses formatted data internally
report__us = generate_currency_report_regional(df_experiment_us, df_final_us_exclusion, df_pg_us, df_sbp_us, 'US', 'Usd')
report__ca = generate_currency_report_regional(df_experiment_ca, df_final_ca_exclusion, df_pg_ca, df_sbp_ca, 'CA', 'Lc')


# %%
# ============================================================================
# CREATE PROGRAM NAME SUMMARY SHEETS
# ============================================================================
print('\n[*] Creating Program Name summary sheets...')

def create_program_summary(df_report, region_name):
    """
    Create summary by Program Name - sum all numeric columns grouped by Program Name
    """
    print(f'[*] Creating {region_name} Program Name summary...')
    
    # Identify numeric columns (excluding Program Name and PRODUCT_LINE) while preserving order
    exclude_cols = ['Program Name', 'PRODUCT_LINE']
    # Get all columns from the original report in order, then filter for numeric ones
    all_cols = df_report.columns.tolist()
    numeric_cols = [col for col in all_cols if col not in exclude_cols and df_report[col].dtype in ['int64', 'float64']]
    
    # Group by Program Name and sum all numeric columns
    program_summary = df_report.groupby('Program Name')[numeric_cols].sum().reset_index()
    
    # Reorder columns to match original report order (Program Name first, then numeric columns in order)
    final_cols = ['Program Name'] + numeric_cols
    program_summary = program_summary[final_cols]
    
    # Round all numeric values to 2 decimal places
    program_summary[numeric_cols] = program_summary[numeric_cols].round(2)
    
    print(f'[*] {region_name} Program Name summary created with {len(program_summary)} program entries')
    return program_summary

# Create program summaries for both US and CA
program_summary_us = create_program_summary(report__us, 'US')
program_summary_ca = create_program_summary(report__ca, 'CA')

# %%
logger.info("Initiating data export to Excel files.")

# Export US Report with Program Summary
with pd.ExcelWriter('Final_Report_USD_US.xlsx', engine='openpyxl', mode='w') as writer:
    report__us.to_excel(writer, sheet_name='USD Report US', index=False)
    program_summary_us.to_excel(writer, sheet_name='Program Summary', index=False)

print('[*] US report exported to Final_Report_USD_US.xlsx')
    
# Export CA Report with Program Summary
with pd.ExcelWriter('Final_Report_Canada_CA.xlsx', engine='openpyxl', mode='w') as writer:
    report__ca.to_excel(writer, sheet_name='LC Report CA', index=False)
    program_summary_ca.to_excel(writer, sheet_name='Program Summary', index=False)

print('[*] CA report exported to Final_Report_Canada_CA.xlsx')

# %%
# NOTE: Attach and Annuity reports are now generated within individual partner reports
# This section has been moved to the comprehensive partner report generation function

# ============================================================================
# DISTY_PARTNERS AND EXCLUSIONS FILTERING FOR US AND CA
# ============================================================================
print('\n--------------------------------------------------------------')
print('[*] Processing Disty_Partners and Exclusions filtering for US and CA...')

def filter_and_export_disty_data(df_data, region_name, currency_col):
    """
    Filter data based on Disty_Partners and Exclusions, then create separate datasets
    for each Disty_Partners entry.
    """
    print(f'\n[*] Processing {region_name} data...')
    
    # Step 1: Filter rows where 'Disty_Partners' is not NaN
    df_disty_filtered = df_data[df_data['Disty_Partners'].notna()]
    print(f'[*] {region_name} - Rows after Disty_Partners filtering: {len(df_disty_filtered)}')

    # Step 2: Apply Scheme_Name filtering logic for compute, storage, and services
    compute_storage_filtered_base = df_disty_filtered[
        df_disty_filtered['Scheme_Name'].str.contains('Compute|Storage', case=False, na=False) &
        (df_disty_filtered['DATA_TYPE'] == 'DS')
    ]

    compute_storage_orders_n3 = df_disty_filtered[
        df_disty_filtered['Scheme_Name'].str.contains('Compute|Storage', case=False, na=False) &
        (df_disty_filtered['DATA_TYPE'].isin(['Orders', 'S4DOR'])) &
        (df_disty_filtered['PRODUCT_LINE'] == 'N3')
    ]

    compute_storage_filtered = pd.concat([compute_storage_filtered_base, compute_storage_orders_n3], ignore_index=True)

    services_filtered = df_disty_filtered[
        df_disty_filtered['Scheme_Name'].str.contains('Services', case=False, na=False) &
        (df_disty_filtered['DATA_TYPE'].isin(['DS', 'Orders', 'S4DOR']))
    ].copy()

    # Apply Services-specific Scheme_Name updates
    for idx in services_filtered.index:
        pn_standalone = services_filtered.loc[idx, 'PN_Standalone']

        if pn_standalone in ['Services Focus standalone', 'Services Standard standalone']:
            continue

        if pd.isna(pn_standalone) or pn_standalone == '':
            common_pn_pl = services_filtered.loc[idx, 'Common_PN_PL']

            if not (pd.isna(common_pn_pl) or common_pn_pl == '') and common_pn_pl == 'Common PL':
                portfolio_mapping_1 = services_filtered.loc[idx, 'Point Next Portfolio Mapping 1'] if 'Point Next Portfolio Mapping 1' in services_filtered.columns else ''
                portfolio_mapping_2 = services_filtered.loc[idx, 'Point Next Portfolio Mapping 2'] if 'Point Next Portfolio Mapping 2' in services_filtered.columns else ''

                if (str(portfolio_mapping_1).startswith('Operational Service') and 
                    str(portfolio_mapping_2).startswith('Complete Care (excl. MS & GL)')):
                    services_filtered.loc[idx, 'Scheme_Name'] = 'ServicesFocus'
                elif (not (pd.isna(portfolio_mapping_1) or portfolio_mapping_1 == '') and 
                      not (pd.isna(portfolio_mapping_2) or portfolio_mapping_2 == '') and 
                      not str(portfolio_mapping_2).startswith('Complete Care (excl. MS & GL)')):
                    services_filtered.loc[idx, 'Scheme_Name'] = 'ServicesStandard'

    df_combined_filtered = pd.concat([compute_storage_filtered, services_filtered], ignore_index=True)
    print(f'[*] {region_name} - Rows after Scheme_Name filtering: {len(df_combined_filtered)}')

    # Step 3: Filter rows where 'Exclusions' is not NaN
    df_exclusions_filtered = df_combined_filtered[df_combined_filtered['Exclusions'].notna()]
    print(f'[*] {region_name} - Rows after Exclusions filtering: {len(df_exclusions_filtered)}')
    
    # Define the required columns
    required_columns = [
        'DISTRIBUTOR_PARTY_ID', 'HPE_SALES_ORDER_NUMBER', 'INVOICE_NUMBER', 
        'Exclusions', 'PRODUCT_LINE', 'RESELLER_PARTY_ID', currency_col,
        'Disty_Partners'  # Include for grouping
    ]
    
    # Filter to only required columns
    df_filtered = df_exclusions_filtered[required_columns].copy()
    
    # Create datasets for each Disty_Partners
    datasets = {}
    for disty_partner in df_filtered['Disty_Partners'].unique():
        partner_data = df_filtered[df_filtered['Disty_Partners'] == disty_partner].copy()
        # Drop the Disty_Partners column
        partner_data = partner_data.drop('Disty_Partners', axis=1)
        datasets[disty_partner] = partner_data
        print(f'[*] {region_name} - Processed Disty_Partner {int(disty_partner)}: {len(partner_data)} rows')

    return datasets

# Process US data
us_datasets = filter_and_export_disty_data(df_exclusions_columns_final_us, 'US', 'NET_TOTAL_USD')

# Process CA data
ca_datasets = filter_and_export_disty_data(df_exclusions_columns_final_ca, 'CA', 'NET_TOTAL_LC')

# ============================================================================
# EXCLUSIONS DATA INTEGRATION
# ============================================================================
print('\n[*] Exclusions data will be integrated into comprehensive partner reports...')
print(f'[*] US partners with exclusions data: {len(us_datasets)} partners')
print(f'[*] CA partners with exclusions data: {len(ca_datasets)} partners')

# %%
# ============================================================================
# SUMMARY REPORT GENERATION FUNCTION
# ============================================================================
def create_summary_report(partner_data, partner_id, region_name, start_date, end_date, days_reporting, days_quarter, partner_report=None):
    """
    Create a Summary report with predefined columns for each partner
    partner_report: DSO DataSheet report containing Program Names and PG/SBP Coverage Sales
    """
    print(f'[*] Creating Summary report for {region_name} Partner {int(partner_id)}...')
    
    # Debug: Print partner_report info
    if partner_report is not None:
        print(f'[DEBUG] Partner report shape: {partner_report.shape}')
        print(f'[DEBUG] Partner report columns: {list(partner_report.columns)}')
        if 'Program Name' in partner_report.columns:
            print(f'[DEBUG] Unique program names: {partner_report["Program Name"].unique()}')
        else:
            print(f'[DEBUG] WARNING: "Program Name" column not found in partner_report')
    else:
        print(f'[DEBUG] partner_report is None')
    
    # Program % mapping dictionary
    program_percentage_mapping = {
        'ComputeStandard_PG': 0.005,
        'StorageStandard_PG': 0.005,
        'ComputeStandard_SBP': 0.055,
        'StorageStandard_SBP': 0.055,
        'ComputeFocus_PG': 0.035,
        'StorageFocus_PG': 0.035,
        'ComputeFocus_SBP': 0.065,
        'StorageFocus_SBP': 0.065,
        'ServicesStandard_PG': 0.02,
        'ServicesStandard_SBP': 0.04,
        'ServicesFocus_PG': 0.04,
        'ServicesFocus_SBP': 0.06
    }
    
    # Initialize summary data lists
    summary_data = []
    
    if partner_report is not None and len(partner_report) > 0:
        # Create Program Summary to get aggregated PG and SBP Coverage Sales by Program Name
        program_summary = partner_report.groupby('Program Name').agg({
            'PG Coverage Sales': 'sum',
            'SBP Coverage Sales': 'sum'
        }).reset_index()
        
        # Extract unique program names from Program Summary
        unique_programs = program_summary['Program Name'].unique()
        
        # Create entries for each program with _PG and _SBP suffixes
        for program_name in unique_programs:
            if pd.notna(program_name) and program_name.strip() != '':
                # Get the aggregated data for this program name from Program Summary
                program_summary_row = program_summary[program_summary['Program Name'] == program_name].iloc[0]
                
                # Extract total PG and SBP coverage sales values from Program Summary
                pg_total_sales = program_summary_row.get('PG Coverage Sales', 0.0)
                sbp_total_sales = program_summary_row.get('SBP Coverage Sales', 0.0)
                
                # Handle NaN values
                if pd.isna(pg_total_sales):
                    pg_total_sales = 0.0
                if pd.isna(sbp_total_sales):
                    sbp_total_sales = 0.0
                
                # Create PG entry
                pg_key = f"{program_name}_PG"
                pg_program_percentage = program_percentage_mapping.get(pg_key, 0.0)
                pg_amount = pg_total_sales * pg_program_percentage
                
                # Debug: Print mapping lookup
                print(f'[DEBUG] PG Mapping - Program: {program_name}, Key: {pg_key}, Percentage: {pg_program_percentage}, Coverage Sales: {pg_total_sales}, Amount: {pg_amount}')
                
                # Determine Incentive Name based on region
                incentive_name = 'US FinBen' if region_name == 'US' else 'CA FinBen'
                
                summary_data.append({
                    'Incentive Name': incentive_name,
                    'New Name': f"{program_name}_PG",
                    'Start Date': start_date,
                    'End Date': end_date,
                    'Amount': pg_amount,
                    'Company Name': '',
                    'Program %': pg_program_percentage,
                    'Days of Reporting': days_reporting,
                    'Days in Quarter': days_quarter,
                    'Program Goal': 0.0,
                    'SBP SDI': 0.0,
                    'Project Attainment': 0.0,
                    'Net QTD Performance': pg_total_sales,
                    'Net Projected Performance': (pg_total_sales / days_reporting) * days_quarter if days_reporting > 0 else 0.0,
                    'Adjusted Amount': 0.0,
                    'Total Rebate Amount': pg_amount,
                    'Projected Rebate Amount': (pg_amount / days_reporting) * days_quarter if days_reporting > 0 else 0.0,
                    'Comments': ''
                })
                
                # Create SBP entry
                sbp_program_percentage = program_percentage_mapping.get(f"{program_name}_SBP", 0.0)
                sbp_amount = sbp_total_sales * sbp_program_percentage
                summary_data.append({
                    'Incentive Name': incentive_name,
                    'New Name': f"{program_name}_SBP",
                    'Start Date': start_date,
                    'End Date': end_date,
                    'Amount': sbp_amount,
                    'Company Name': '',
                    'Program %': sbp_program_percentage,
                    'Days of Reporting': days_reporting,
                    'Days in Quarter': days_quarter,
                    'Program Goal': 0.0,
                    'SBP SDI': 0.0,
                    'Project Attainment': 0.0,
                    'Net QTD Performance': sbp_total_sales,
                    'Net Projected Performance': (sbp_total_sales / days_reporting) * days_quarter if days_reporting > 0 else 0.0,
                    'Adjusted Amount': 0.0,
                    'Total Rebate Amount': sbp_amount,
                    'Projected Rebate Amount': (sbp_amount / days_reporting) * days_quarter if days_reporting > 0 else 0.0,
                    'Comments': ''
                })
    
    # If no data available, create empty structure
    if not summary_data:
        # Determine Incentive Name based on region for empty structure
        incentive_name = 'US FinBen' if region_name == 'US' else 'CA FinBen'
        
        summary_data = [{
            'Incentive Name': incentive_name,
            'New Name': '',
            'Start Date': '',
            'End Date': '',
            'Amount': 0.0,
            'Company Name': '',
            'Program %': 0.0,
            'Days of Reporting': 0,
            'Days in Quarter': 0,
            'Program Goal': 0.0,
            'SBP SDI': 0.0,
            'Project Attainment': 0.0,
            'Net QTD Performance': 0.0,
            'Net Projected Performance': 0.0,
            'Adjusted Amount': 0.0,
            'Total Rebate Amount': 0.0,
            'Projected Rebate Amount': 0.0,
            'Comments': ''
        }]
    
    # Create DataFrame from summary data
    summary_df = pd.DataFrame(summary_data)
    
    print(f'[*] Summary report created for {region_name} Partner {int(partner_id)} with {len(summary_df)} rows')
    return summary_df

# ============================================================================
# SUMMARY REPORT DETAILS SETUP
# ============================================================================
print('\n--------------------------------------------------------------')
print('[*] Setting up Summary report details...')
print('--------------------------------------------------------------')

# Calculate fiscal quarter start and end dates
today = datetime.today()
current_month = today.month
current_year = today.year

# Define fiscal quarter boundaries (fiscal year starts in November)
if current_month in [11, 12]:  # Q1: Nov, Dec (current year)
    # Q1 starts November 1st of current year, ends January 31st of next year
    start_date = datetime(current_year, 11, 1).strftime('%Y-%m-%d')
    end_date = datetime(current_year + 1, 1, 31).strftime('%Y-%m-%d')
elif current_month == 1:  # Q1: Jan (next year)
    # Q1 starts November 1st of previous year, ends January 31st of current year
    start_date = datetime(current_year - 1, 11, 1).strftime('%Y-%m-%d')
    end_date = datetime(current_year, 1, 31).strftime('%Y-%m-%d')
elif current_month in [2, 3, 4]:  # Q2: Feb, Mar, Apr
    # Q2 starts February 1st, ends April 30th
    start_date = datetime(current_year, 2, 1).strftime('%Y-%m-%d')
    end_date = datetime(current_year, 4, 30).strftime('%Y-%m-%d')
elif current_month in [5, 6, 7]:  # Q3: May, Jun, Jul
    # Q3 starts May 1st, ends July 31st
    start_date = datetime(current_year, 5, 1).strftime('%Y-%m-%d')
    end_date = datetime(current_year, 7, 31).strftime('%Y-%m-%d')
else:  # Q4: Aug, Sep, Oct
    # Q4 starts August 1st, ends October 31st
    start_date = datetime(current_year, 8, 1).strftime('%Y-%m-%d')
    end_date = datetime(current_year, 10, 31).strftime('%Y-%m-%d')

# Calculate region-specific days_quarter from days reporting files
# Get the maximum 'Days of Reporting' value from each region's days reporting file
try:
    days_quarter_ca = df_days_ca['Days of Reporting'].max() if 'Days of Reporting' in df_days_ca.columns and len(df_days_ca) > 0 else 0 # type: ignore
    days_quarter_us = df_days_us['Days of Reporting'].max() if 'Days of Reporting' in df_days_us.columns and len(df_days_us) > 0 else 0 # type: ignore
    print(f'[*] Days Quarter CA (from max days reporting): {days_quarter_ca}')
    print(f'[*] Days Quarter US (from max days reporting): {days_quarter_us}')
except Exception as e:
    print(f'[!] Error calculating days_quarter from reporting files: {e}')
    # Fallback to calculated quarter days
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
    days_quarter_fallback = (end_date_obj - start_date_obj).days + 1
    days_quarter_ca = days_quarter_fallback
    days_quarter_us = days_quarter_fallback
    print(f'[*] Using fallback days_quarter: {days_quarter_fallback}')

print(f'\n[*] Summary report details:')
print(f'    Start Date: {start_date}')
print(f'    End Date: {end_date}')
print(f'    Days of Reporting (CA): {days_reporting_ca}')
print(f'    Days of Reporting (US): {days_reporting_us}')
print(f'    Days in Quarter (CA): {days_quarter_ca}')
print(f'    Days in Quarter (US): {days_quarter_us}')

# ============================================================================
# COMPREHENSIVE DISTY_PARTNERS REPORT GENERATION
# ============================================================================
print('\n--------------------------------------------------------------')
print('[*] Starting comprehensive Disty_Partners report generation...')

# Create output directories

os.makedirs('US_partners_report', exist_ok=True)
os.makedirs('Canada_partners_report', exist_ok=True)

def process_disty_partner_comprehensive(df_final_data, df_exclusions_final, region_name, currency_type, folder_name):
    """
    Process complete monthly reports for each Disty_Partners entry, including separate PG and SBP summaries.
    """
    print(f'\n[*] Processing {region_name} comprehensive Disty_Partners reports...')

    # Get unique Disty_Partners (excluding NaN)
    unique_partners = df_final_data['Disty_Partners'].dropna().unique()
    print(f'[*] {region_name} - Found {len(unique_partners)} unique Disty_Partners: {[int(p) for p in unique_partners]}')

    for partner in unique_partners:
        print(f'\n[*] Processing {region_name} Disty_Partner {int(partner)}...')

        # Filter data for this specific partner
        partner_final_data = df_final_data[df_final_data['Disty_Partners'] == partner].copy()
        partner_exclusion_data = df_exclusions_final[df_exclusions_final['Disty_Partners'] == partner].copy()

        if len(partner_final_data) == 0:
            print(f'[!] {region_name} Partner {int(partner)} - No data found, skipping...')
            continue

        print(f'[*] {region_name} Partner {int(partner)} - Processing {len(partner_final_data)} rows')

        # Create PG/SBP separation for this partner
        partner_pg = partner_final_data[partner_final_data['PG_Exclusions'] == 'PG']
        partner_sbp = partner_final_data[partner_final_data['PG_Exclusions'] == 'SBP']

        # Prepare monthly data for this partner
        try:
            partner_monthly_data = prepare_monthly_data(partner_final_data.copy(), f'{region_name} Partner {int(partner)}')

            if len(partner_monthly_data) == 0:
                print(f'[!] {region_name} Partner {int(partner)} - No monthly data after filtering')
                continue

            # Generate currency report for this partner
            partner_report = generate_currency_report_regional(
                partner_monthly_data,
                partner_exclusion_data,
                partner_pg,
                partner_sbp,
                f'{region_name} Partner {int(partner)}',
                currency_type
            )

            # Create program summary for this partner
            partner_program_summary = create_program_summary(partner_report, f'{region_name} Partner {int(partner)}')

            # Create Summary report for this partner with the partner's DSO DataSheet as input
            # Use region-specific days_reporting values
            if region_name == 'US':
                days_reporting_value = days_reporting_us
                days_quarter_value = days_quarter_us
            else:  # CA
                days_reporting_value = days_reporting_ca
                days_quarter_value = days_quarter_ca
            
            partner_summary_report = create_summary_report(partner_final_data, partner, region_name, start_date, end_date, days_reporting_value, days_quarter_value, partner_report)

            # Get exclusions dataset for this partner
            partner_exclusions_dataset = None
            if region_name == 'US' and partner in us_datasets:
                partner_exclusions_dataset = us_datasets[partner]
            elif region_name == 'CA' and partner in ca_datasets:
                partner_exclusions_dataset = ca_datasets[partner]

            # Export to Excel file
            filename = f'{folder_name}/Disty_Partner_{int(partner)}_Report.xlsx'

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Main report
                partner_report.to_excel(writer, sheet_name=f'DSO DataSheet', index=False)

                # Program summary
                partner_program_summary.to_excel(writer, sheet_name='Program Summary', index=False)

                # Summary report
                partner_summary_report.to_excel(writer, sheet_name='Summary', index=False)

                # Exclusions dataset (if available)
                if partner_exclusions_dataset is not None and len(partner_exclusions_dataset) > 0:
                    partner_exclusions_dataset.to_excel(writer, sheet_name='Exclusions Data', index=False)

                # Add Attach and Annuity reports for specific US partners
                print(f'[DEBUG] Checking Attach/Annuity generation for Partner {int(partner)}')
                print(f'[DEBUG] Region: {region_name}, Partner in target list: {partner in [1000939629, 1001810197]}')
                print(f'[DEBUG] Partner type: {type(partner)}, Partner value: {partner}')
                
                if region_name == 'US' and partner in [1000939629, 1001810197]:
                    print(f'[*] Adding Attach and Annuity reports for US Partner {int(partner)}...')
                    print(f'[DEBUG] partner_monthly_data shape: {partner_monthly_data.shape}')
                    print(f'[DEBUG] partner_monthly_data columns: {list(partner_monthly_data.columns)}')

                    # Define the columns to extract for both reports
                    common_columns = [
                        'fiscal_quarter', 'DISTRIBUTOR_PARTY_ID', 'Distributor Party Name',
                        'RESELLER_PARTY_ID', 'Reseller Party Name', 'PRODUCT_LINE',
                        'Product Number', 'NET_TOTAL_USD'
                    ]
                    
                    print(f'[DEBUG] Required common_columns: {common_columns}')

                    # Check if required columns exist in partner data
                    available_columns = [col for col in common_columns if col in partner_monthly_data.columns]
                    missing_columns = [col for col in common_columns if col not in partner_monthly_data.columns]
                    
                    print(f'[DEBUG] Available columns from common_columns: {available_columns}')
                    print(f'[DEBUG] Missing columns from common_columns: {missing_columns}')
                    print(f'[DEBUG] Invoice Number in columns: {"Invoice Number" in partner_monthly_data.columns}')
                    print(f'[DEBUG] Hpe Sales Order Number in columns: {"Hpe Sales Order Number" in partner_monthly_data.columns}')

                    if len(available_columns) == len(common_columns):
                        print(f'[DEBUG] All required columns are available, proceeding with report generation...')
                        
                        # Report 1: Attach - with Invoice Number
                        if 'INVOICE_NUMBER' in partner_monthly_data.columns:
                            attach_columns = available_columns + ['INVOICE_NUMBER']
                            attach_report = partner_monthly_data[attach_columns].copy()
                            # Rename fiscal_quarter to Fiscal Quarter for consistency
                            attach_report = attach_report.rename(columns={'fiscal_quarter': 'Fiscal Quarter'})
                            attach_report.to_excel(writer, sheet_name='Attach', index=False)
                            print(f'[*] Attach report added for Partner {int(partner)} with {len(attach_report)} rows')
                        else:
                            print(f'[!] Cannot create Attach report - Invoice Number column missing')

                        # Report 2: Annuity - with Hpe Sales Order Number
                        if 'HPE_SALES_ORDER_NUMBER' in partner_monthly_data.columns:
                            annuity_columns = available_columns + ['HPE_SALES_ORDER_NUMBER']
                            annuity_report = partner_monthly_data[annuity_columns].copy()
                            # Rename fiscal_quarter to Fiscal Quarter for consistency
                            annuity_report = annuity_report.rename(columns={'fiscal_quarter': 'Fiscal Quarter'})
                            annuity_report.to_excel(writer, sheet_name='Annuity', index=False)
                            print(f'[*] Annuity report added for Partner {int(partner)} with {len(annuity_report)} rows')
                        else:
                            print(f'[!] Cannot create Annuity report - Hpe Sales Order Number column missing')
                    else:
                        print(f'[!] Missing required columns for Attach/Annuity reports for Partner {int(partner)}')
                        print(f'[!] Available: {available_columns}')
                        print(f'[!] Required: {common_columns}')
                        print(f'[!] Missing: {missing_columns}')
                else:
                    if region_name == 'US':
                        print(f'[DEBUG] US Partner {int(partner)} not in target list [1000939629, 1001810197] - skipping Attach/Annuity')
                    else:
                        print(f'[DEBUG] Non-US region ({region_name}) - skipping Attach/Annuity')

            print(f'[*] {region_name} Partner {int(partner)} - Report exported to {filename}')

        except Exception as e:
            print(f'[!] Error processing {region_name} Partner {int(partner)}: {e}')
            continue

# Process US partners
process_disty_partner_comprehensive(
    df_final_us, df_final_us_exclusion, 'US', 'Usd', 'US_partners_report'
)

# Process CA partners
process_disty_partner_comprehensive(
    df_final_ca, df_final_ca_exclusion, 'CA', 'Lc', 'Canada_partners_report'
)

# %%
# Note: Exclusions data is now integrated into comprehensive partner reports
print('\n[*] Exclusions data integration completed.')
print('[*] Each partner report includes:')
print('    - Final Report sheet (main analysis)')
print('    - Program Summary sheet (aggregated by program)')
print('    - Exclusions Data sheet (detailed exclusions data when available)')

print('\n' + '='*60)
print('COMPREHENSIVE FLASH REPORT GENERATION COMPLETED')
print('='*60)
print('[*] Final output files:')
print('    - Final_Report_USD_US.xlsx (US data with USD currency)')
print('    - Final_Report_Canada_CA.xlsx (CA data with LC currency)')
print('    - US file includes Attach and Annuity sheets for specific Disty_Partners')
print('    - Both files include Program Summary sheets')
print(f'    - US_partners_report/ folder with {len(us_datasets)} partner reports')
print(f'    - Canada_partners_report/ folder with {len(ca_datasets)} partner reports')
print('[*] Script execution completed successfully!')

# End of script

# # %%
# ============================================================================
# OPERATION 2: READ ALL SHEETS FROM NEW EXCEL FILE
# ============================================================================
print('\n' + '='*60)
print('OPERATION 2: READ ALL SHEETS FROM NEW EXCEL FILE')
print('='*60)

def read_all_sheets_from_excel(file_path):
    """
    Returns:
    dict: Dictionary with sheet names as keys and DataFrames as values
    """
    print(f'[*] Operation 2: Reading all sheets from - {file_path}')
    
    # Validate file existence
    if not os.path.isfile(file_path):
        print(f"[!] Error: File '{file_path}' does not exist.")
        return None
    
    try:
        # Read all sheets from the Excel file
        all_sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
        
        print(f'[*] Successfully loaded {len(all_sheets)} sheets from {os.path.basename(file_path)}')
        
        # Display information about each sheet
        for sheet_name, df in all_sheets.items():
            print(f'[*] Sheet "{sheet_name}": {df.shape[0]} rows × {df.shape[1]} columns')
            print(f'    Columns: {list(df.columns)}')
        
        return all_sheets
        
    except Exception as e:
        print(f"[!] Error reading Excel file: {e}")
        return None

# ============================================================================
# OPERATION 2: EXECUTE WITH COMMAND LINE ARGUMENT
# ============================================================================
print('\n[*] Operation 2: Processing new Excel file from command line argument')
print(f'[*] New Excel file path: {s3_raw_data_path}')

# Execute Operation 2 with the provided file path
all_sheets_from_new_file = read_all_sheets_from_excel(s3_raw_data_path)

if all_sheets_from_new_file is not None:
    print(f'\n[*] Operation 2: Successfully processed {len(all_sheets_from_new_file)} sheets')
    print('[*] Available sheets:')
    for sheet_name in all_sheets_from_new_file.keys():
        print(f'    - {sheet_name}')
    
    # Access all 5 sheets
    sheet1_df = all_sheets_from_new_file[list(all_sheets_from_new_file.keys())[0]]  # US data
    sheet2_df = all_sheets_from_new_file[list(all_sheets_from_new_file.keys())[1]]  # CA data
    sheet3_df = all_sheets_from_new_file[list(all_sheets_from_new_file.keys())[2]]  # Eligible sales file
    sheet4_df = all_sheets_from_new_file[list(all_sheets_from_new_file.keys())[3]]  # US reference sheet sheet
    sheet5_df = all_sheets_from_new_file[list(all_sheets_from_new_file.keys())[4]]  # Canada reference sheet sheet
    
    # For backward compatibility, maintain the original variable names
    us_data_df = sheet1_df
    ca_data_df = sheet2_df
    eligible_sales = sheet3_df
    us_reference_sheet = sheet4_df
    ca_reference_sheet = sheet5_df
    
    # Check for required columns in eligible_sales sheet
    print(f'[*] Checking required columns in eligible_sales sheet...')
    print(f'[*] Available columns in eligible_sales: {list(eligible_sales.columns)}')
    
    required_eligible_sales_columns = ["Saas eligible PL's", 'US_Loc Id', 'CA_Loc Id']
    missing_eligible_sales_columns = []
    
    for col in required_eligible_sales_columns:
        if col not in eligible_sales.columns:
            missing_eligible_sales_columns.append(col)
            # Try to find similar column names
            similar_cols = get_close_matches(col, eligible_sales.columns, n=3, cutoff=0.6)
            if similar_cols:
                print(f'[!] Warning: Column "{col}" not found in eligible_sales. Similar columns: {similar_cols}')
            else:
                print(f'[!] Warning: Column "{col}" not found in eligible_sales. No similar columns found.')
    
    if missing_eligible_sales_columns:
        print(f'[!] Error: Missing required columns in eligible_sales: {missing_eligible_sales_columns}')
        print('[!] Some operations may be skipped due to missing eligible_sales columns.')
    
    # Check for required columns in US data sheet
    required_us_columns = ['PRODUCTLINE_ID', 'REPORTING_SELLER_ID', 'BUYER_PARTNER_ID']
    missing_us_columns = []
    
    print(f'[*] Checking required columns in US data sheet...')
    print(f'[*] Available columns in US data: {list(us_data_df.columns)}')
    
    for col in required_us_columns:
        if col not in us_data_df.columns:
            missing_us_columns.append(col)
            # Try to find similar column names
            similar_cols = get_close_matches(col, us_data_df.columns, n=3, cutoff=0.6)
            if similar_cols:
                print(f'[!] Warning: Column "{col}" not found in US data. Similar columns: {similar_cols}')
            else:
                print(f'[!] Warning: Column "{col}" not found in US data. No similar columns found.')
    
    if missing_us_columns:
        print(f'[!] Error: Missing required columns in US data: {missing_us_columns}')
        print('[!] Skipping US data processing due to missing columns.')
        us_data_df = pd.DataFrame()  # Create empty DataFrame
    else:
        # Operation on the first sheet (US based data)
        # Add a new column 'EG BU' (only if eligible_sales has required column)
        if "Saas eligible PL's" in eligible_sales.columns:
            us_data_df['EG BU'] = us_data_df['PRODUCTLINE_ID'].apply(
                lambda pl_id: pl_id if pl_id in eligible_sales["Saas eligible PL's"].values else None
            )
        else:
            print('[!] Warning: Cannot create EG BU column - "Saas eligible PL\'s" column missing in eligible_sales')
            us_data_df['EG BU'] = None
    
        print(f'[*] Updated US data sheet with new column "EG BU": {us_data_df.shape[0]} rows × {us_data_df.shape[1]} columns')
        
        # Add a second column 'Disty Partner' (only if eligible_sales has required column)
        if 'US_Loc Id' in eligible_sales.columns:
            us_data_df['Disty Partner'] = us_data_df['REPORTING_SELLER_ID'].apply(
                lambda seller_id: seller_id if seller_id in eligible_sales['US_Loc Id'].values else None
            )
        else:
            print('[!] Warning: Cannot create Disty Partner column - "US_Loc Id" column missing in eligible_sales')
            us_data_df['Disty Partner'] = None
        
        print(f'[*] Updated US data sheet with new column "Disty Partner": {us_data_df.shape[0]} rows × {us_data_df.shape[1]} columns')
        
        # Add a third column 'US_PG Exclusions'
        us_data_df['US_PG Exclusions'] = us_data_df['BUYER_PARTNER_ID'].apply(
            lambda buyer_id: 'PG' if buyer_id in us_reference_sheet['PG Exclusion Eligible List_Party ID'].values else 'SBP'
        )
        
        print(f'[*] Updated US data sheet with new column "US_PG Exclusions": {us_data_df.shape[0]} rows × {us_data_df.shape[1]} columns')
        
        # Add the fourth column 'US_Exclusions'
        # Create a mapping Series from 'Exclusion_Party ID' to 'Exclusion_Level' in the US reference sheet using groupby
        exclusion_mapping = us_reference_sheet.groupby('Exclusion_Party ID')['Exclusion_Level'].first()
        
        # Populate 'US_Exclusions' column by mapping 'BUYER_PARTNER_ID' via the exclusion mapping; missing matches fill with None
        us_data_df['US_Exclusions'] = us_data_df['BUYER_PARTNER_ID'].map(exclusion_mapping)
        
        print(f'[*] Updated US data sheet with new column "US_Exclusions": {us_data_df.shape[0]} rows × {us_data_df.shape[1]} columns')

        # Filter for non-empty and non-NaN "EG BU" and "Disty Partner"
        initial_row_count = us_data_df.shape[0]
        us_data_df = us_data_df[us_data_df['EG BU'].notna() & us_data_df['Disty Partner'].notna()]
        print(f'[*] Filtered "EG BU" and "Disty Partner": from {initial_row_count} to {us_data_df.shape[0]} rows')

        # Filter for "US_Exclusions" being NaN or empty
        initial_row_count = us_data_df.shape[0]
        us_data_df = us_data_df[us_data_df['US_Exclusions'].isna() | (us_data_df['US_Exclusions'] == '')]
        print(f'[*] Filtered "US_Exclusions": from {initial_row_count} to {us_data_df.shape[0]} rows')

    # Operation on the second sheet (CA based data)
    ca_data_df = sheet2_df
    
    # Check for required columns in CA data sheet
    required_ca_columns = ['PRODUCTLINE_ID', 'REPORTING_SELLER_ID', 'BUYER_PARTNER_ID']
    missing_ca_columns = []
    
    print(f'[*] Checking required columns in CA data sheet...')
    print(f'[*] Available columns in CA data: {list(ca_data_df.columns)}')
    
    for col in required_ca_columns:
        if col not in ca_data_df.columns:
            missing_ca_columns.append(col)
            # Try to find similar column names
            similar_cols = get_close_matches(col, ca_data_df.columns, n=3, cutoff=0.6)
            if similar_cols:
                print(f'[!] Warning: Column "{col}" not found in CA data. Similar columns: {similar_cols}')
            else:
                print(f'[!] Warning: Column "{col}" not found in CA data. No similar columns found.')
    
    if missing_ca_columns:
        print(f'[!] Error: Missing required columns in CA data: {missing_ca_columns}')
        print('[!] Skipping CA data processing due to missing columns.')
        ca_data_df = pd.DataFrame()  # Create empty DataFrame
    else:
        # Add a new column 'EG BU' (only if eligible_sales has required column)
        if "Saas eligible PL's" in eligible_sales.columns:
            ca_data_df['EG BU'] = ca_data_df['PRODUCTLINE_ID'].apply(
                lambda pl_id: pl_id if pl_id in eligible_sales["Saas eligible PL's"].values else None
            )
        else:
            print('[!] Warning: Cannot create EG BU column - "Saas eligible PL\'s" column missing in eligible_sales')
            ca_data_df['EG BU'] = None
    
        print(f'[*] Updated CA data sheet with new column "EG BU": {ca_data_df.shape[0]} rows × {ca_data_df.shape[1]} columns')
        
        # Add a second column 'Disty Partner' (only if eligible_sales has required column)
        if 'CA_Loc Id' in eligible_sales.columns:
            ca_data_df['Disty Partner'] = ca_data_df['REPORTING_SELLER_ID'].apply(
                lambda seller_id: seller_id if seller_id in eligible_sales['CA_Loc Id'].values else None
            )
        else:
            print('[!] Warning: Cannot create Disty Partner column - "CA_Loc Id" column missing in eligible_sales')
            ca_data_df['Disty Partner'] = None
        
        print(f'[*] Updated CA data sheet with new column "Disty Partner": {ca_data_df.shape[0]} rows × {ca_data_df.shape[1]} columns')
        
        # Add a third column 'CA_PG Exclusions'
        ca_data_df['CA_PG Exclusions'] = ca_data_df['BUYER_PARTNER_ID'].apply(
            lambda buyer_id: 'PG' if buyer_id in ca_reference_sheet['PG Exclusion Eligible List_Party ID'].values else 'SBP'
        )
        
        print(f'[*] Updated CA data sheet with new column "CA_PG Exclusions": {ca_data_df.shape[0]} rows × {ca_data_df.shape[1]} columns')
        
        # Add the fourth column 'CA_Exclusions'
        # Create a mapping Series from 'Exclusion_Party ID' to 'Exclusion_Level' in the CA reference sheet using groupby
        ca_exclusion_mapping = ca_reference_sheet.groupby('Exclusion_Party ID')['Exclusion_Level'].first()
        
        # Populate 'CA_Exclusions' column by mapping 'BUYER_PARTNER_ID' via the exclusion mapping; missing matches fill with None
        ca_data_df['CA_Exclusions'] = ca_data_df['BUYER_PARTNER_ID'].map(ca_exclusion_mapping)
        
        print(f'[*] Updated CA data sheet with new column "CA_Exclusions": {ca_data_df.shape[0]} rows × {ca_data_df.shape[1]} columns')

        # Filter for non-empty and non-NaN "EG BU" and "Disty Partner" in CA data
        initial_row_count = ca_data_df.shape[0]
        ca_data_df = ca_data_df[ca_data_df['EG BU'].notna() & ca_data_df['Disty Partner'].notna()]
        print(f'[*] Filtered CA "EG BU" and "Disty Partner": from {initial_row_count} to {ca_data_df.shape[0]} rows')

        # Filter for "CA_Exclusions" being NaN or empty
        initial_row_count = ca_data_df.shape[0]
        ca_data_df = ca_data_df[ca_data_df['CA_Exclusions'].isna() | (ca_data_df['CA_Exclusions'] == '')]
        print(f'[*] Filtered CA "CA_Exclusions": from {initial_row_count} to {ca_data_df.shape[0]} rows')

    # Add 'Scheme Name' column for US data based on US_PG Exclusions values (only if US data was processed)
    if not us_data_df.empty and 'US_PG Exclusions' in us_data_df.columns:
        us_data_df['Scheme Name'] = us_data_df['US_PG Exclusions'].apply(
            lambda pg_value: f'ComputeFocus_{pg_value}' if pd.notna(pg_value) else None
        )
        print(f'[*] Added "Scheme Name" column to US data: {us_data_df.shape[0]} rows × {us_data_df.shape[1]} columns')
    else:
        print('[!] Skipping US "Scheme Name" column - US data not available or missing columns')
    
    # Add 'Scheme Name' column for CA data based on CA_PG Exclusions values (only if CA data was processed)
    if not ca_data_df.empty and 'CA_PG Exclusions' in ca_data_df.columns:
        ca_data_df['Scheme Name'] = ca_data_df['CA_PG Exclusions'].apply(
            lambda pg_value: f'ComputeFocus_{pg_value}' if pd.notna(pg_value) else None
        )
        print(f'[*] Added "Scheme Name" column to CA data: {ca_data_df.shape[0]} rows × {ca_data_df.shape[1]} columns')
    else:
        print('[!] Skipping CA "Scheme Name" column - CA data not available or missing columns')

    # Calculate TCV summaries for US and CA data
    print('\n[*] Calculating TCV summaries for Disty Partner and Scheme Name combinations...')
    
    # Initialize variables to avoid unbound variable warnings
    us_tcv_summary = pd.DataFrame()
    ca_tcv_summary = pd.DataFrame()
    
    # For US Data - sum TCV_NET_EXTENDED_AMOUNT by Disty Partner and Scheme Name
    if not us_data_df.empty and 'TCV_NET_EXTENDED_AMOUNT' in us_data_df.columns and 'Disty Partner' in us_data_df.columns and 'Scheme Name' in us_data_df.columns:
        us_tcv_summary = us_data_df.groupby(['Disty Partner', 'Scheme Name'])['TCV_NET_EXTENDED_AMOUNT'].sum().reset_index()
        print(f'[*] US TCV Summary calculated: {len(us_tcv_summary)} combinations')
        print(us_tcv_summary)
    else:
        print('[WARNING] Cannot calculate US TCV Summary - missing required columns or empty data')
    
    # For CA Data - sum TCV_NET_EXTENDED_AMOUNT_LC by Disty Partner and Scheme Name
    if not ca_data_df.empty and 'TCV_NET_EXTENDED_AMOUNT_LC' in ca_data_df.columns and 'Disty Partner' in ca_data_df.columns and 'Scheme Name' in ca_data_df.columns:
        ca_tcv_summary = ca_data_df.groupby(['Disty Partner', 'Scheme Name'])['TCV_NET_EXTENDED_AMOUNT_LC'].sum().reset_index()
        print(f'[*] CA TCV Summary calculated: {len(ca_tcv_summary)} combinations')
        print(ca_tcv_summary)
    else:
        print('[WARNING] Cannot calculate CA TCV Summary - missing required columns or empty data')
    
    # Update individual partner files with TCV amounts
    print('\n[*] Updating individual partner files with TCV amounts...')
    
    def update_partner_file_with_tcv(partner_id, tcv_summary, region_folder, currency_col):
        """Update partner file's DSO DataSheet with new ComputeFocus row containing PG and SBP Coverage Sales from TCV data"""
        file_path = f'{region_folder}/Disty_Partner_{int(partner_id)}_Report.xlsx'
        
        if not os.path.exists(file_path):
            print(f'[!] Partner file not found: {file_path}')
            return
        
        try:
            # Read the existing file
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # Read all sheets first
                existing_file = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
                
                # Process DSO DataSheet if it exists
                if 'DSO DataSheet' in existing_file:
                    dso_df = existing_file['DSO DataSheet'].copy()
                    
                    # Get TCV data for this partner
                    partner_tcv_data = tcv_summary[tcv_summary['Disty Partner'] == partner_id]
                    
                    if len(partner_tcv_data) > 0:
                        print(f'[*] Updating Partner {int(partner_id)} DSO DataSheet with TCV data:')
                        
                        # Aggregate TCV amounts by scheme type (PG vs SBP)
                        pg_total = 0.0
                        sbp_total = 0.0
                        
                        for _, tcv_row in partner_tcv_data.iterrows():
                            scheme_name = tcv_row['Scheme Name']
                            tcv_amount = tcv_row[currency_col]
                            
                            if 'ComputeFocus_PG' in scheme_name:
                                pg_total += tcv_amount
                                print(f'    - PG: {scheme_name} = {tcv_amount:,.2f}')
                            elif 'ComputeFocus_SBP' in scheme_name:
                                sbp_total += tcv_amount
                                print(f'    - SBP: {scheme_name} = {tcv_amount:,.2f}')
                        
                        # Only add row if we have TCV data
                        if pg_total > 0 or sbp_total > 0:
                            # Create new row for ComputeFocus with aggregated PG/SBP amounts
                            new_row = {}
                            
                            # Get all columns from existing DSO DataSheet
                            for col in dso_df.columns:
                                if col == 'Program Name':
                                    new_row[col] = 'ComputeFocus'
                                elif col == 'PRODUCT_LINE':
                                    new_row[col] = 'S3'  # Assuming S3 since Operation 2 processes S3 data
                                elif col == 'PG Coverage Sales':
                                    new_row[col] = pg_total
                                elif col == 'SBP Coverage Sales':
                                    new_row[col] = sbp_total
                                else:
                                    # Set all other columns to 0 or empty
                                    if dso_df[col].dtype in ['int64', 'float64']:
                                        new_row[col] = 0.0
                                    else:
                                        new_row[col] = ''
                            
                            # Add the new row to the dataframe
                            new_row_df = pd.DataFrame([new_row])
                            dso_df = pd.concat([dso_df, new_row_df], ignore_index=True)
                            
                            print(f'    - Added ComputeFocus row: PG Coverage Sales = {pg_total:,.2f}, SBP Coverage Sales = {sbp_total:,.2f}')
                    
                    # Write updated DSO DataSheet
                    dso_df.to_excel(writer, sheet_name='DSO DataSheet', index=False)
                    
                    # Copy all other sheets unchanged
                    for sheet_name, sheet_df in existing_file.items():
                        if sheet_name != 'DSO DataSheet':
                            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    print(f'[*] Updated partner file: {file_path}')
                else:
                    print(f'[!] DSO DataSheet not found in: {file_path}')
                    
        except Exception as e:
            print(f'[!] Error updating partner file {file_path}: {e}')
    
    # Update US partner files (only if US TCV summary was successfully calculated)
    if not us_tcv_summary.empty and 'Disty Partner' in us_tcv_summary.columns:
        print(f'[*] Updating {len(us_tcv_summary["Disty Partner"].unique())} US partner files...')
        for partner_id in us_tcv_summary['Disty Partner'].unique():
            update_partner_file_with_tcv(partner_id, us_tcv_summary, 'US_partners_report', 'TCV_NET_EXTENDED_AMOUNT')
    else:
        print('[!] Skipping US partner file updates - no TCV summary data available')
    
    # Update CA partner files (only if CA TCV summary was successfully calculated)
    if not ca_tcv_summary.empty and 'Disty Partner' in ca_tcv_summary.columns:
        print(f'[*] Updating {len(ca_tcv_summary["Disty Partner"].unique())} CA partner files...')
        for partner_id in ca_tcv_summary['Disty Partner'].unique():
            update_partner_file_with_tcv(partner_id, ca_tcv_summary, 'Canada_partners_report', 'TCV_NET_EXTENDED_AMOUNT_LC')
    else:
        print('[!] Skipping CA partner file updates - no TCV summary data available')
    
    print('\n[*] Partner file updates completed!')
    
    print('\n[*] Operation 2: All sheets have been loaded into memory')
    print('[*] Use all_sheets_from_new_file["sheet_name"] to access specific sheets')
else:
    print('[!] Operation 2: Failed to read the new Excel file')

print('\n[*] Operation 2 completed successfully!')

# %%
# ============================================================================
# OPERATION 3: READ REBATE FILE WITH SPECIFIC SHEETS
# ============================================================================
print('\n' + '='*60)
print('OPERATION 3: READ REBATE FILE')
print('='*60)

def read_rebate_sheets(file_path):
    """
    Read specific sheets 'RebateSummary' and 'RebateDetails' from rebate file
    Returns:
    tuple: (rebate_summary_df, rebate_details_df)
    """
    print(f'[*] Operation 3: Reading rebate sheets from - {file_path}')
    
    # Validate file existence
    if not os.path.isfile(file_path):
        print(f"[!] Error: File '{file_path}' does not exist.")
        return None, None
    
    try:
        # Read specific sheets
        rebate_summary_df = None
        rebate_details_df = None
        
        # Try to read RebateSummary sheet
        try:
            rebate_summary_df = pd.read_excel(file_path, sheet_name='RebateSummary', engine='openpyxl')
            print(f'[*] RebateSummary sheet loaded: {rebate_summary_df.shape[0]} rows × {rebate_summary_df.shape[1]} columns')
            print(f'    Columns: {list(rebate_summary_df.columns)}')
        except Exception as e:
            print(f'[!] Warning: Could not read RebateSummary sheet: {e}')
        
        # Try to read RebateDetails sheet
        try:
            rebate_details_df = pd.read_excel(file_path, sheet_name='RebateDetails', engine='openpyxl')
            print(f'[*] RebateDetails sheet loaded: {rebate_details_df.shape[0]} rows × {rebate_details_df.shape[1]} columns')
            print(f'    Columns: {list(rebate_details_df.columns)}')
        except Exception as e:
            print(f'[!] Warning: Could not read RebateDetails sheet: {e}')
        
        return rebate_summary_df, rebate_details_df
        
    except Exception as e:
        print(f"[!] Error reading rebate file: {e}")
        return None, None

# ============================================================================
# OPERATION 3: EXECUTE WITH COMMAND LINE ARGUMENT
# ============================================================================
print('\n[*] Operation 3: Processing rebate file from command line argument')
print(f'[*] Rebate file path: {rebate_file_path}')

# Execute Operation 3 with the provided file path
rebate_summary_data, rebate_details_data = read_rebate_sheets(rebate_file_path)

if rebate_summary_data is not None or rebate_details_data is not None:
    print(f'\n[*] Operation 3: Successfully processed rebate file')
    
    if rebate_summary_data is not None:
        print(f'[*] RebateSummary data available with {len(rebate_summary_data)} rows')
        # Store for further processing if needed
        rebate_summary_df = rebate_summary_data
    else:
        print('[*] RebateSummary data not available')
        rebate_summary_df = None
    
    if rebate_details_data is not None:
        print(f'[*] RebateDetails data available with {len(rebate_details_data)} rows')
        # Store for further processing if needed
        rebate_details_df = rebate_details_data
    else:
        print('[*] RebateDetails data not available')
        rebate_details_df = None
    
    print('\n[*] Operation 3: Rebate data loaded into memory')
    print('[*] Use rebate_summary_df and rebate_details_df to access the data')
else:
    print('[!] Operation 3: Failed to read any sheets from the rebate file')
    rebate_summary_df = None
    rebate_details_df = None

print('\n[*] Operation 3 completed successfully!')

# ============================================================================
# UPDATE COMPANY NAMES IN PARTNER FILES (AFTER OPERATION 3)
# ============================================================================
# Create Company Name mapping from eligible_sales sheet
print('\n[*] Creating Company Name mapping from eligible_sales sheet...')
us_company_name_mapping = {}
ca_company_name_mapping = {}

# Check if eligible_sales variable exists (from Operation 2)
if 'eligible_sales' in locals() and eligible_sales is not None: # type: ignore
    print('[*] Eligible sales data available for company name mapping')
    
    # Check if eligible_sales has the required columns and create mappings
    if 'US_Loc Id' in eligible_sales.columns and 'US_RS Company Name' in eligible_sales.columns: # type: ignore
        # Create mapping for US Loc Id to US_RS Company Name
        us_data = eligible_sales[eligible_sales['US_Loc Id'].notna() & eligible_sales['US_RS Company Name'].notna()] # type: ignore
        us_company_name_mapping = dict(zip(us_data['US_Loc Id'].astype(int), us_data['US_RS Company Name'].astype(str)))
        print(f'[*] Created US Loc Id to Company Name mapping: {len(us_company_name_mapping)} entries')
    else:
        missing_cols = []
        if 'US_Loc Id' not in eligible_sales.columns: # type: ignore
            missing_cols.append('US_Loc Id')
        if 'US_RS Company Name' not in eligible_sales.columns: # type: ignore
            missing_cols.append('US_RS Company Name')
        print(f'[!] Warning: Missing columns for US mapping: {missing_cols}')
    
    if 'CA_Loc Id' in eligible_sales.columns and 'CA_RS Company Name' in eligible_sales.columns: # type: ignore
        # Create mapping for CA Loc Id to CA_RS Company Name
        ca_data = eligible_sales[eligible_sales['CA_Loc Id'].notna() & eligible_sales['CA_RS Company Name'].notna()] # type: ignore
        ca_company_name_mapping = dict(zip(ca_data['CA_Loc Id'].astype(int), ca_data['CA_RS Company Name'].astype(str)))
        print(f'[*] Created CA Loc Id to Company Name mapping: {len(ca_company_name_mapping)} entries')
    else:
        missing_cols = []
        if 'CA_Loc Id' not in eligible_sales.columns: # type: ignore
            missing_cols.append('CA_Loc Id')
        if 'CA_RS Company Name' not in eligible_sales.columns: # type: ignore
            missing_cols.append('CA_RS Company Name')
        print(f'[!] Warning: Missing columns for CA mapping: {missing_cols}')
    
    print(f'[*] Available columns in eligible_sales: {list(eligible_sales.columns)}') # type: ignore
else:
    print('[!] Warning: Eligible sales data not available from Operation 2')
    print('[!] Company name mapping will be skipped')

# Function to update Company Name in Summary and Summary2 sheets
def update_summary_sheet_company_name(partner_id, region_folder, company_mapping, region_name):
    """
    Update Company Name column in Summary sheet immediately after creation
    """
    file_path = f'{region_folder}/{partner_id}_partner_report.xlsx'
    
    if not os.path.exists(file_path):
        print(f'[!] Partner file not found: {file_path}')
        return
    
    try:
        # Read the Excel file
        with pd.ExcelFile(file_path) as xls:
            # Read all sheets into dictionary
            all_sheets = {}
            for sheet_name in xls.sheet_names:
                all_sheets[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
        
        # Update Summary sheet if it exists and has Company Name column
        if 'Summary' in all_sheets and 'Company Name' in all_sheets['Summary'].columns:
            company_name = company_mapping.get(partner_id, f'Partner {partner_id}')
            all_sheets['Summary']['Company Name'] = company_name
            
            # Write back to Excel
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in all_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            print(f'[*] Updated {region_name} Partner {partner_id} Summary sheet with company name: {company_name}')
        else:
            print(f'[*] No Summary sheet or Company Name column found for {region_name} Partner {partner_id}')
    
    except Exception as e:
        print(f'[!] Error updating Summary sheet company name for partner {partner_id}: {e}')

def update_summary2_sheet_company_name(partner_id, region_folder, company_mapping, region_name):
    """
    Update Company Name column in Summary2 sheet after its creation in Operation 3
    """
    file_path = f'{region_folder}/Disty_Partner_{int(partner_id)}_Report.xlsx'
    
    if not os.path.exists(file_path):
        print(f'[!] Partner file not found: {file_path}')
        return
    
    try:
        # Read the Excel file
        with pd.ExcelFile(file_path) as xls:
            # Read all sheets into dictionary
            all_sheets = {}
            for sheet_name in xls.sheet_names:
                all_sheets[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
        
        # Update Summary2 sheet if it exists and has Company Name column
        if 'Summary2' in all_sheets and 'Company Name' in all_sheets['Summary2'].columns:
            company_name = company_mapping.get(partner_id, f'Partner {partner_id}')
            all_sheets['Summary2']['Company Name'] = company_name
            
            # Write back to Excel
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in all_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            print(f'[*] Updated {region_name} Partner {partner_id} Summary2 sheet with company name: {company_name}')
        else:
            print(f'[*] No Summary2 sheet or Company Name column found for {region_name} Partner {partner_id}')
    
    except Exception as e:
        print(f'[!] Error updating Summary2 sheet company name for partner {partner_id}: {e}')

def update_partner_company_names(partner_id, region_folder, company_mapping, region_name):
    """Update Company Name column in Summary and Summary2 sheets"""
    file_path = f'{region_folder}/Disty_Partner_{int(partner_id)}_Report.xlsx'

    if not os.path.exists(file_path):
        print(f'[!] Partner file not found: {file_path}')
        return

    # Get company name for this partner
    company_name = company_mapping.get(int(partner_id), '')

    if not company_name:
        print(f'[!] No company name found for {region_name} Partner {int(partner_id)}')
        return

    try:
        # Read all sheets from the existing file
        existing_file = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')

        sheets_updated = []

        # Update Summary sheet if it exists
        if 'Summary' in existing_file and 'Company Name' in existing_file['Summary'].columns:
            summary_df = existing_file['Summary'].copy()
            summary_df['Company Name'] = company_name
            existing_file['Summary'] = summary_df
            sheets_updated.append('Summary')

        # Update Summary2 sheet if it exists
        if 'Summary2' in existing_file and 'Company Name' in existing_file['Summary2'].columns:
            summary2_df = existing_file['Summary2'].copy()
            summary2_df['Company Name'] = company_name
            existing_file['Summary2'] = summary2_df
            sheets_updated.append('Summary2')

        # Write back to Excel if any sheets were updated
        if sheets_updated:
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                for sheet_name, sheet_df in existing_file.items():
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

            print(f'[*] Updated Company Name for {region_name} Partner {int(partner_id)} ({company_name}) in sheets: {", ".join(sheets_updated)}')
        else:
            print(f'[!] No Summary or Summary2 sheets with Company Name column found for {region_name} Partner {int(partner_id)}')

    except Exception as e:
        print(f'[!] Error updating company name for partner file {file_path}: {e}')

# Update company names in partner files
print('\n[*] Updating Company Names in partner files...')

# Update US partner files
if not us_tcv_summary.empty and us_company_name_mapping: # type: ignore
    print('[*] Updating US partner files with Company Names...')
    for partner_id in us_tcv_summary['Disty Partner'].unique(): # type: ignore
        update_partner_company_names(partner_id, 'US_partners_report', us_company_name_mapping, 'US')
elif us_tcv_summary.empty: # type: ignore
    print('[!] No US TCV summary data available')
elif not us_company_name_mapping:
    print('[!] No US company name mapping available')

# Update CA partner files
if not ca_tcv_summary.empty and ca_company_name_mapping: # type: ignore
    print('[*] Updating CA partner files with Company Names...')
    for partner_id in ca_tcv_summary['Disty Partner'].unique(): # type: ignore
        update_partner_company_names(partner_id, 'Canada_partners_report', ca_company_name_mapping, 'CA')
elif ca_tcv_summary.empty: # type: ignore
    print('[!] No CA TCV summary data available')
elif not ca_company_name_mapping:
    print('[!] No CA company name mapping available')

print('[*] Company Name updates completed!')

# %%
# ============================================================================
# OPERATION 3: PROCESS REBATE SUMMARY DATA FOR SUMMARY2 SHEETS
# ============================================================================
if rebate_summary_df is not None:
    print('\n' + '='*60)
    print('OPERATION 3: PROCESS REBATE SUMMARY DATA FOR SUMMARY2 SHEETS')
    print('='*60)
    
    print(f'[*] Processing RebateSummary data with {len(rebate_summary_df)} rows')
    print(f'[*] Available columns: {list(rebate_summary_df.columns)}')
    
    # Step 1: Filter by COUNTRY (US and Canada)
    print('\n[*] Step 1: Filtering by COUNTRY (US and Canada)...')
    us_rebate_data = pd.DataFrame()
    ca_rebate_data = pd.DataFrame()
    
    if 'COUNTRY' in rebate_summary_df.columns:
        us_rebate_data = rebate_summary_df[rebate_summary_df['COUNTRY'].str.upper() == 'US'].copy()
        ca_rebate_data = rebate_summary_df[rebate_summary_df['COUNTRY'].str.upper().isin(['CA', 'CANADA'])].copy()
        
        print(f'[*] US rebate data: {len(us_rebate_data)} rows')
        print(f'[*] CA rebate data: {len(ca_rebate_data)} rows')
    else:
        print('[!] Warning: COUNTRY column not found in RebateSummary')
    
    # Step 2: Filter by PARTNER_TYPE = 'Distributor'
    print('\n[*] Step 2: Filtering by PARTNER_TYPE = Distributor...')
    if 'PARTNER_TYPE' in rebate_summary_df.columns:
        us_rebate_data = us_rebate_data[us_rebate_data['PARTNER_TYPE'].str.upper() == 'DISTRIBUTOR']
        ca_rebate_data = ca_rebate_data[ca_rebate_data['PARTNER_TYPE'].str.upper() == 'DISTRIBUTOR']
        
        print(f'[*] US rebate data after PARTNER_TYPE filter: {len(us_rebate_data)} rows')
        print(f'[*] CA rebate data after PARTNER_TYPE filter: {len(ca_rebate_data)} rows')
    else:
        print('[!] Warning: PARTNER_TYPE column not found, skipping this filter')
    
    # Step 3: Eliminate values starting with 'Lar' and 'NA A&G1'
    print('\n[*] Step 3: Eliminating values starting with "Lar" and containing "LAC"...')
    
    def filter_exclusions(df, region_name):
        """Filter out rows with values starting with 'Lar' or containing 'lac'"""
        if len(df) == 0:
            return df
        
        initial_count = len(df)
        
        # Check all string columns for exclusion patterns
        for col in df.select_dtypes(include=['object']).columns:
            # Remove rows starting with 'Lar' (case insensitive)
            df = df[~df[col].str.upper().str.startswith('LAR', na=False)]
            # # Remove rows containing 'NA A&G1'
            # df = df[~df[col].str.contains('NA A&G1', case=False, na=False)]
            # Remove rows containing 'lac' (case insensitive)
            df = df[~df[col].str.contains('lac', case=False, na=False)]
        
        print(f'    {region_name}: Filtered from {initial_count} to {len(df)} rows')
        return df
    
    us_rebate_data = filter_exclusions(us_rebate_data, 'US')
    ca_rebate_data = filter_exclusions(ca_rebate_data, 'CA')
    
    # Step 4: Filter SCHEME column for L1, L2, L3 values and sum FINAL_REBATE
    print('\n[*] Step 4: Processing SCHEME column for L1, L2, L3 values...')
    
    def process_scheme_data(df, region_name):
        """Process scheme data for L1, L2, L3 and sum FINAL_REBATE by PARTY_ID"""
        if len(df) == 0:
            print(f'[*] {region_name}: No data to process')
            return pd.DataFrame()
        
        required_columns = ['SCHEME', 'FINAL_REBATE', 'PARTY_ID']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            print(f'[!] {region_name}: Missing required columns: {missing_columns}')
            print(f'[!] Available columns: {list(df.columns)}')
            return pd.DataFrame()
        
        # Filter for schemes containing L1, L2, L3, or starting with 'aaS' (case sensitive)
        scheme_filter = (df['SCHEME'].str.contains('L1|L2|L3', case=False, na=False) | 
                        df['SCHEME'].str.startswith('aaS', na=False))
        filtered_df = df[scheme_filter].copy()
        
        print(f'[*] {region_name}: Found {len(filtered_df)} rows with L1/L2/L3 schemes or schemes starting with "aaS"')
        
        if len(filtered_df) == 0:
            return pd.DataFrame()
        
        # Group by PARTY_ID and SCHEME, sum FINAL_REBATE
        grouped_data = filtered_df.groupby(['PARTY_ID', 'SCHEME'])['FINAL_REBATE'].sum().reset_index()
        
        print(f'[*] {region_name}: Grouped into {len(grouped_data)} PARTY_ID-SCHEME combinations')
        return grouped_data
    
    us_scheme_data = process_scheme_data(us_rebate_data, 'US')
    ca_scheme_data = process_scheme_data(ca_rebate_data, 'CA')
    
    print(f'[*] US scheme data: {len(us_scheme_data)} rows')
    print(f'[*] CA scheme data: {len(ca_scheme_data)} rows')
    
    # Step 5: Match with Disty Partner IDs from Operation 2 and update partner files
    print('\n[*] Step 5: Matching with Disty Partner IDs and updating partner files...')
    
    # Define scheme to dollar amount mapping for Program % column
    scheme_program_mapping = {
        'Compute Expansion L1': 700,
        'Compute Expansion L2': 750,
        'Compute Expansion L3': 800,
        'Services Expansion L1': 250,
        'Services Expansion L2': 600,
        'Storage Expansion L1': 650,
        'Storage Expansion L2': 700,
        'Storage Expansion L3': 1000
    }
    
    def update_partner_summary2_sheet(partner_id, scheme_data, region_folder, region_name):
        """Update partner file with Summary2 sheet containing L1, L2, L3, and aaS scheme data"""
        file_path = f'{region_folder}/Disty_Partner_{int(partner_id)}_Report.xlsx'
        
        if not os.path.exists(file_path):
            print(f'[!] Partner file not found: {file_path}')
            return
        
        # Filter scheme data for this specific partner
        partner_scheme_data = scheme_data[scheme_data['PARTY_ID'] == partner_id]
        
        if len(partner_scheme_data) == 0:
            print(f'[*] No scheme data found for {region_name} Partner {int(partner_id)}')
            return
        
        try:
            # Read the existing file to get Summary sheet structure
            existing_file = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
            
            if 'Summary' not in existing_file:
                print(f'[!] Summary sheet not found in {file_path}')
                return
            
            # Get the structure from Summary sheet
            summary_structure = existing_file['Summary'].copy()
            
            # Create Summary2 data with same structure
            summary2_data = []
            
            for _, row in partner_scheme_data.iterrows():
                # Create a new row with Summary sheet structure
                new_row = {col: '' for col in summary_structure.columns}
                
                # Populate key fields
                new_row['New Name'] = row['SCHEME']
                new_row['Amount'] = row['FINAL_REBATE']
                new_row['Total Rebate Amount'] = row['FINAL_REBATE']
                
                # Set Net QTD Performance equal to Amount
                new_row['Net QTD Performance'] = row['FINAL_REBATE']
                
                # Populate Incentive Name based on region
                if 'Incentive Name' in new_row:
                    new_row['Incentive Name'] = f'{region_name} FinBen'
                
                # Populate Program % column with mapped dollar value
                scheme_name = row['SCHEME']
                program_value = scheme_program_mapping.get(scheme_name, '')
                if 'Program %' in new_row:
                    new_row['Program %'] = str(program_value) if program_value != '' else ''
                
                # Copy some fields from the first row of Summary if available
                days_of_reporting = 0
                days_in_quarter = 0
                if len(summary_structure) > 0:
                    first_summary_row = summary_structure.iloc[0]
                    for col in ['Start Date', 'End Date', 'Days of Reporting', 'Days in Quarter']:
                        if col in new_row:
                            new_row[col] = first_summary_row.get(col, '')
                            if col == 'Days of Reporting':
                                try:
                                    days_of_reporting = float(first_summary_row.get(col, 0))
                                except (ValueError, TypeError):
                                    days_of_reporting = 0
                            elif col == 'Days in Quarter':
                                try:
                                    days_in_quarter = float(first_summary_row.get(col, 0))
                                except (ValueError, TypeError):
                                    days_in_quarter = 0
                
                # Calculate projections if we have valid days values
                if days_of_reporting > 0 and days_in_quarter > 0:
                    # Net Projected Performance = (Net QTD Performance / Days of Reporting) * Days in Quarter
                    net_projected_performance = (row['FINAL_REBATE'] / days_of_reporting) * days_in_quarter
                    new_row['Net Projected Performance'] = net_projected_performance
                    
                    # Projected Rebate Amount = (Total Rebate Amount / Days of Reporting) * Days in Quarter
                    projected_rebate_amount = (row['FINAL_REBATE'] / days_of_reporting) * days_in_quarter
                    new_row['Projected Rebate Amount'] = projected_rebate_amount
                else:
                    # Set to empty if we can't calculate
                    new_row['Net Projected Performance'] = ''
                    new_row['Projected Rebate Amount'] = ''
                
                summary2_data.append(new_row)
            
            # Create Summary2 DataFrame
            summary2_df = pd.DataFrame(summary2_data)
            
            # Create Summary2 as a separate sheet (not appending to Summary)
            existing_file['Summary2'] = summary2_df
            
            # Write back to Excel with separate Summary2 sheet
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                # Write all sheets including the new Summary2 sheet
                for sheet_name, sheet_df in existing_file.items():
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Add highlighting to the 'New Name' column in Summary2
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            
            # Load the workbook and access the Summary2 sheet
            wb = load_workbook(file_path)
            if 'Summary2' in wb.sheetnames:
                ws = wb['Summary2']
                
                green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                
                # Find the 'New Name' column index by reading the header row
                new_name_col_index = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == 'New Name':
                        new_name_col_index = col
                        break
                
                # Apply green background to all cells in the 'New Name' column
                if new_name_col_index is not None:
                    for row in range(1, ws.max_row + 1):
                        ws.cell(row=row, column=new_name_col_index).fill = green_fill
                
                # Save the workbook
                wb.save(file_path)
            
            print(f'[*] Created Summary2 sheet for {region_name} Partner {int(partner_id)} with {len(summary2_df)} rows from rebate data')
            for _, row in partner_scheme_data.iterrows():
                print(f'    - {row["SCHEME"]}: {row["FINAL_REBATE"]:,.2f}')
            
            # Update Company Name in Summary2 sheet immediately after creation
            if region_name == 'US' and us_company_name_mapping:
                update_summary2_sheet_company_name(partner_id, region_folder, us_company_name_mapping, region_name)
            elif region_name == 'CA' and ca_company_name_mapping:
                update_summary2_sheet_company_name(partner_id, region_folder, ca_company_name_mapping, region_name)
            
        except Exception as e:
            print(f'[!] Error updating partner file {file_path}: {e}')
    
    # Process US partners
    if len(us_scheme_data) > 0 and not us_tcv_summary.empty: # type: ignore
        print('\n[*] Processing US partners...')
        us_partner_ids = us_tcv_summary['Disty Partner'].unique() # type: ignore
        
        processed_count = 0
        for partner_id in us_partner_ids:
            if partner_id in us_scheme_data['PARTY_ID'].values:
                update_partner_summary2_sheet(partner_id, us_scheme_data, 'US_partners_report', 'US')
                processed_count += 1
        
        print(f'[*] Processed {processed_count} US partners with scheme data')
    else:
        print('[*] No US scheme data or partner IDs available')
    
    # Process CA partners
    if len(ca_scheme_data) > 0 and not ca_tcv_summary.empty: # type: ignore
        print('\n[*] Processing CA partners...')
        ca_partner_ids = ca_tcv_summary['Disty Partner'].unique() # type: ignore
        
        processed_count = 0
        for partner_id in ca_partner_ids:
            if partner_id in ca_scheme_data['PARTY_ID'].values:
                update_partner_summary2_sheet(partner_id, ca_scheme_data, 'Canada_partners_report', 'CA')
                processed_count += 1
        
        print(f'[*] Processed {processed_count} CA partners with scheme data')
    else:
        print('[*] No CA scheme data or partner IDs available')
    
    print('\n[*] Summary2 sheet updates completed!')
    print('[*] Company Names updated immediately after each Summary2 sheet creation.')

else:
    print('[!] No RebateSummary data loaded - skipping Summary2 processing')

logger.info("All operations completed successfully. Flash Report process has finished.")
print('\n[*] Operation 3 data processing completed successfully!')